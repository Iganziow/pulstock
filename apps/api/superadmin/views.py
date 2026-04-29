"""
superadmin/views.py
===================
Panel de superadministración de la plataforma Pulstock.
Todos los endpoints requieren is_superuser=True.
"""

from api.utils import safe_int
from django.db import models, transaction, IntegrityError
from django.db.models import Count, Sum, Q, F, Avg, FloatField
from django.db.models.fields.json import KeyTextTransform
from django.db.models.functions import TruncDate, TruncMonth, Cast
from django.utils import timezone
from django.shortcuts import get_object_or_404
from datetime import timedelta

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status

from core.models import Tenant, User
from core.permissions import IsSuperAdmin
from billing.models import Plan, Subscription, Invoice
from stores.models import Store

import logging
logger = logging.getLogger(__name__)

PERMS = [IsAuthenticated, IsSuperAdmin]


# ─────────────────────────────────────────────────────────────
# DASHBOARD KPIs
# ─────────────────────────────────────────────────────────────
class PlatformStatsView(APIView):
    """KPIs globales de la plataforma."""
    permission_classes = PERMS

    def get(self, request):
        now = timezone.now()
        month_ago = now - timedelta(days=30)

        total_tenants = Tenant.objects.count()
        active_tenants = Tenant.objects.filter(is_active=True).count()
        total_users = User.objects.filter(is_active=True, is_superuser=False).count()
        total_stores = Store.objects.filter(is_active=True).count()

        # Suscripciones por estado
        sub_stats = dict(
            Subscription.objects.values_list("status")
            .annotate(c=Count("id"))
            .values_list("status", "c")
        )

        # Ingresos últimos 30 días
        revenue_30d = Invoice.objects.filter(
            status=Invoice.Status.PAID,
            paid_at__gte=month_ago,
        ).aggregate(total=Sum("amount_clp"))["total"] or 0

        # Ingresos totales
        revenue_total = Invoice.objects.filter(
            status=Invoice.Status.PAID,
        ).aggregate(total=Sum("amount_clp"))["total"] or 0

        # Tenants nuevos últimos 30 días
        new_tenants_30d = Tenant.objects.filter(created_at__gte=month_ago).count()

        # MRR (Monthly Recurring Revenue)
        mrr = Subscription.objects.filter(
            status__in=["active", "trialing", "past_due"],
        ).aggregate(
            mrr=Sum("plan__price_clp")
        )["mrr"] or 0

        # Tenants con pagos pendientes
        past_due = Subscription.objects.filter(status="past_due").count()
        suspended = Subscription.objects.filter(status="suspended").count()

        # Distribución por plan
        plan_dist = list(
            Subscription.objects.filter(
                status__in=["active", "trialing", "past_due"],
            ).values(
                plan_name=F("plan__name"),
                plan_key=F("plan__key"),
            ).annotate(count=Count("id")).order_by("-count")
        )

        return Response({
            "total_tenants": total_tenants,
            "active_tenants": active_tenants,
            "new_tenants_30d": new_tenants_30d,
            "total_users": total_users,
            "total_stores": total_stores,
            "subscription_stats": sub_stats,
            "revenue_30d": revenue_30d,
            "revenue_total": revenue_total,
            "mrr": mrr,
            "past_due": past_due,
            "suspended": suspended,
            "plan_distribution": plan_dist,
        })


# ─────────────────────────────────────────────────────────────
# REVENUE CHART (últimos 12 meses)
# ─────────────────────────────────────────────────────────────
class RevenueChartView(APIView):
    permission_classes = PERMS

    def get(self, request):
        months = safe_int(request.query_params.get("months"), 12)
        since = timezone.now() - timedelta(days=months * 30)

        data = list(
            Invoice.objects.filter(
                status=Invoice.Status.PAID,
                paid_at__gte=since,
            ).annotate(
                month=TruncMonth("paid_at")
            ).values("month").annotate(
                total=Sum("amount_clp"),
                count=Count("id"),
            ).order_by("month")
        )

        for d in data:
            d["month"] = d["month"].strftime("%Y-%m")

        return Response(data)


# ─────────────────────────────────────────────────────────────
# TENANT LIST + DETAIL
# ─────────────────────────────────────────────────────────────
class TenantListView(APIView):
    permission_classes = PERMS

    def get(self, request):
        qs = Tenant.objects.all().order_by("-created_at")

        # Filtros
        q = request.query_params.get("q", "").strip()
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(slug__icontains=q) | Q(email__icontains=q))

        status_filter = request.query_params.get("status")
        if status_filter == "active":
            qs = qs.filter(is_active=True)
        elif status_filter == "inactive":
            qs = qs.filter(is_active=False)

        plan_filter = request.query_params.get("plan")
        if plan_filter:
            qs = qs.filter(subscription__plan__key=plan_filter)

        sub_status = request.query_params.get("sub_status")
        if sub_status:
            qs = qs.filter(subscription__status=sub_status)

        # Annotate
        qs = qs.annotate(
            user_count=Count("users", filter=Q(users__is_active=True, users__is_superuser=False)),
            store_count=Count("stores", filter=Q(stores__is_active=True), distinct=True),
        )

        # Paginación simple
        page = safe_int(request.query_params.get("page"), 1)
        page_size = safe_int(request.query_params.get("page_size"), 25)
        total = qs.count()
        offset = (page - 1) * page_size
        tenants = qs[offset:offset + page_size]

        results = []
        # Prefetch subscriptions
        tenant_ids = [t.id for t in tenants]
        subs = {
            s.tenant_id: s
            for s in Subscription.objects.filter(tenant_id__in=tenant_ids).select_related("plan")
        }

        for t in tenants:
            sub = subs.get(t.id)
            results.append({
                "id": t.id,
                "name": t.name,
                "slug": t.slug,
                "email": t.email,
                "business_type": getattr(t, "business_type", "other") or "other",
                "is_active": t.is_active,
                "created_at": t.created_at.isoformat(),
                "user_count": t.user_count,
                "store_count": t.store_count,
                "subscription": {
                    "status": sub.status if sub else None,
                    "plan_name": sub.plan.name if sub else None,
                    "plan_key": sub.plan.key if sub else None,
                    "price_clp": sub.plan.price_clp if sub else 0,
                    "current_period_end": sub.current_period_end.isoformat() if sub and sub.current_period_end else None,
                    "has_card": bool(sub.card_last4) if sub else False,
                } if sub else None,
            })

        return Response({
            "results": results,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size,
        })

    def post(self, request):
        """Crear tenant desde superadmin con owner, store y warehouse.

        Toda la creación va dentro de un atomic() para que si algo falla
        (ej. owner_email ya existe, o IntegrityError de slug por race),
        no queden objetos huérfanos (Tenant sin Store, etc.).
        """
        from django.utils.text import slugify
        from django.db import transaction, IntegrityError
        from stores.models import Store
        from core.models import Warehouse

        data = request.data
        name = (data.get("name") or "").strip()
        if not name:
            return Response({"detail": "Nombre del negocio es requerido."}, status=400)

        # Validación temprana de owner email (antes del atomic) para devolver
        # 400 limpio en vez de IntegrityError genérico.
        owner_email = data.get("owner_email", "").strip()
        owner_password = data.get("owner_password", "").strip()
        if owner_email and owner_password:
            if User.objects.filter(username=owner_email).exists():
                return Response({"detail": f"El usuario {owner_email} ya existe."}, status=400)

        # Construir slug único — el while no es bulletproof contra race
        # entre 2 superadmin (caso muy raro pero posible). Si la creación
        # falla con IntegrityError lo capturamos abajo y retry una vez.
        base_slug = slugify(name)
        slug = base_slug
        counter = 1
        while Tenant.objects.filter(slug=slug).exists():
            slug = f"{base_slug}-{counter}"
            counter += 1

        try:
            with transaction.atomic():
                tenant = Tenant.objects.create(
                    name=name,
                    slug=slug,
                    business_type=data.get("business_type", "retail"),
                    email=data.get("email", ""),
                    phone=data.get("phone", ""),
                    rut=data.get("rut", ""),
                    legal_name=data.get("legal_name", ""),
                    giro=data.get("giro", ""),
                    address=data.get("address", ""),
                    city=data.get("city", ""),
                    comuna=data.get("comuna", ""),
                )

                store_name = data.get("store_name", "Local Principal")
                store = Store.objects.create(tenant=tenant, name=store_name, is_active=True)
                warehouse = Warehouse.objects.create(
                    tenant=tenant, store=store,
                    name=data.get("warehouse_name", "Bodega Principal"),
                )
                tenant.default_warehouse = warehouse
                tenant.save(update_fields=["default_warehouse"])

                owner_data = {}
                if owner_email and owner_password:
                    owner = User.objects.create_user(
                        username=owner_email,
                        email=owner_email,
                        password=owner_password,
                        first_name=data.get("owner_first_name", ""),
                        last_name=data.get("owner_last_name", ""),
                        tenant=tenant,
                        role="owner",
                        active_store=store,
                    )
                    owner_data = {"id": owner.id, "email": owner.email}

                sub_data = None
                plan_key = data.get("plan_key")
                if plan_key:
                    try:
                        plan = Plan.objects.get(key=plan_key, is_active=True)
                        sub, _ = Subscription.objects.get_or_create(
                            tenant=tenant,
                            defaults={
                                "plan": plan,
                                "status": data.get("sub_status", "active"),
                            },
                        )
                        sub_data = {"plan": plan.name, "status": sub.status}
                    except Plan.DoesNotExist:
                        pass
        except IntegrityError as e:
            # Race condition rara: 2 superadmin creando con el mismo slug
            # exacto a la vez. El atomic hace rollback de todo. Devolvemos
            # 409 conflict para que el frontend pueda reintentar con otro nombre.
            logger.warning(
                "Superadmin tenant create IntegrityError (probable slug race): %s — name=%s",
                e, name,
            )
            return Response(
                {"detail": "El nombre o slug del negocio ya existe. Intenta con otro nombre."},
                status=status.HTTP_409_CONFLICT,
            )

        logger.info("Superadmin %s created tenant %d (%s)", request.user.email, tenant.id, name)

        return Response({
            "id": tenant.id,
            "name": tenant.name,
            "slug": tenant.slug,
            "business_type": tenant.business_type,
            "owner": owner_data,
            "subscription": sub_data,
        }, status=201)


class TenantDetailView(APIView):
    permission_classes = PERMS

    def get(self, request, tenant_id):
        t = get_object_or_404(Tenant, pk=tenant_id)

        users = list(
            User.objects.filter(tenant=t, is_superuser=False).values(
                "id", "username", "email", "first_name", "last_name",
                "role", "is_active", "last_login", "date_joined",
            ).order_by("-date_joined")
        )

        stores = list(
            Store.objects.filter(tenant=t).values(
                "id", "name", "is_active",
            )
        )

        sub = None
        try:
            s = Subscription.objects.select_related("plan").get(tenant=t)
            sub = {
                "id": s.id,
                "status": s.status,
                "plan_key": s.plan.key,
                "plan_name": s.plan.name,
                "price_clp": s.plan.price_clp,
                "trial_ends_at": s.trial_ends_at.isoformat() if s.trial_ends_at else None,
                "current_period_start": s.current_period_start.isoformat() if s.current_period_start else None,
                "current_period_end": s.current_period_end.isoformat() if s.current_period_end else None,
                "has_card": bool(s.card_last4),
                "card_brand": s.card_brand,
                "card_last4": s.card_last4,
                "payment_retry_count": s.payment_retry_count,
            }
        except Subscription.DoesNotExist:
            pass

        invoices = list(
            Invoice.objects.filter(subscription__tenant=t).values(
                "id", "status", "amount_clp", "period_start", "period_end",
                "paid_at", "created_at",
            ).order_by("-created_at")[:20]
        )

        # Conteos rápidos
        from catalog.models import Product
        from sales.models import Sale
        product_count = Product.objects.filter(tenant=t).count()
        sale_count = Sale.objects.filter(tenant=t).count()

        return Response({
            "id": t.id,
            "name": t.name,
            "slug": t.slug,
            "email": t.email,
            "phone": t.phone,
            "rut": t.rut,
            "legal_name": t.legal_name,
            "business_type": getattr(t, "business_type", "other") or "other",
            "is_active": t.is_active,
            "internal_notes": t.internal_notes,
            "created_at": t.created_at.isoformat(),
            "users": users,
            "stores": stores,
            "subscription": sub,
            "invoices": invoices,
            "stats": {
                "products": product_count,
                "sales": sale_count,
            },
        })

    def patch(self, request, tenant_id):
        """Modificar tenant — soporta todos los campos editables."""
        t = get_object_or_404(Tenant, pk=tenant_id)
        changed = []

        # Campos de texto simples
        text_fields = ["name", "email", "phone", "rut", "legal_name", "giro",
                       "address", "city", "comuna", "website", "logo_url", "primary_color"]
        for field in text_fields:
            if field in request.data:
                setattr(t, field, request.data[field])
                changed.append(field)

        if "is_active" in request.data:
            t.is_active = bool(request.data["is_active"])
            changed.append("is_active")

        if "business_type" in request.data:
            valid_types = [c[0] for c in Tenant.BUSINESS_TYPE_CHOICES]
            bt = request.data["business_type"]
            if bt in valid_types:
                t.business_type = bt
                changed.append("business_type")

        if changed:
            t.save(update_fields=changed)
            logger.info("Superadmin %s updated tenant %d: %s", request.user.email, t.id, changed)

        return Response({"ok": True, "id": t.id, "is_active": t.is_active, "name": t.name})

    def delete(self, request, tenant_id):
        """Eliminar tenant y todos sus datos asociados (orden respeta PROTECT FKs)."""
        from sales.models import Sale, SaleLine, SalePayment
        from purchases.models import Purchase, PurchaseLine, PurchaseInvoiceLine, Supplier
        from inventory.models import StockItem, StockMove, StockTransfer, StockTransferLine
        from catalog.models import Barcode, RecipeLine, Recipe, Product, Category, Unit
        from forecast.models import (
            SuggestionOutcome, ForecastAccuracy, SuggestionLine,
            PurchaseSuggestion, Forecast, ForecastModel, CategoryDemandProfile,
            DailySales, Holiday,
        )
        from tables.models import OpenOrderLine, OpenOrder, Table
        from caja.models import CashMovement, CashSession, CashRegister
        from billing.models import CheckoutSession, Invoice, PaymentAttempt, Subscription
        from stores.models import Store
        from core.models import Warehouse

        t = get_object_or_404(Tenant, pk=tenant_id)
        tenant_name = t.name
        user_count = User.objects.filter(tenant=t).count()

        with transaction.atomic():
            # Null out PROTECT FKs that would block deletion
            Tenant.objects.filter(pk=t.pk).update(default_warehouse=None)

            # Delete in dependency order (leaves before branches)
            SalePayment.objects.filter(tenant=t).delete()
            SaleLine.objects.filter(tenant=t).delete()
            Sale.objects.filter(tenant=t).delete()
            OpenOrderLine.objects.filter(tenant=t).delete()
            OpenOrder.objects.filter(tenant=t).delete()
            Table.objects.filter(tenant=t).delete()
            CashMovement.objects.filter(tenant=t).delete()
            CashSession.objects.filter(tenant=t).delete()
            CashRegister.objects.filter(tenant=t).delete()
            PurchaseInvoiceLine.objects.filter(tenant=t).delete()
            PurchaseLine.objects.filter(tenant=t).delete()
            Purchase.objects.filter(tenant=t).delete()
            Supplier.objects.filter(tenant=t).delete()
            SuggestionOutcome.objects.filter(tenant=t).delete()
            ForecastAccuracy.objects.filter(tenant=t).delete()
            SuggestionLine.objects.filter(suggestion__tenant=t).delete()
            PurchaseSuggestion.objects.filter(tenant=t).delete()
            Forecast.objects.filter(tenant=t).delete()
            ForecastModel.objects.filter(tenant=t).delete()
            CategoryDemandProfile.objects.filter(tenant=t).delete()
            Holiday.objects.filter(tenant=t).delete()
            DailySales.objects.filter(tenant=t).delete()
            StockTransferLine.objects.filter(tenant=t).delete()
            StockTransfer.objects.filter(tenant=t).delete()
            StockMove.objects.filter(tenant=t).delete()
            StockItem.objects.filter(tenant=t).delete()
            Barcode.objects.filter(tenant=t).delete()
            RecipeLine.objects.filter(tenant=t).delete()
            Recipe.objects.filter(tenant=t).delete()
            Product.objects.filter(tenant=t).delete()
            Category.objects.filter(tenant=t).update(parent=None)
            Category.objects.filter(tenant=t).delete()
            Unit.objects.filter(tenant=t).delete()
            PaymentAttempt.objects.filter(invoice__subscription__tenant=t).delete()
            CheckoutSession.objects.filter(tenant=t).delete()
            Invoice.objects.filter(subscription__tenant=t).delete()
            Subscription.objects.filter(tenant=t).delete()
            # Null active_store for any user referencing this tenant's stores
            User.objects.filter(active_store__tenant=t).update(active_store=None)
            Warehouse.objects.filter(tenant=t).delete()
            Store.objects.filter(tenant=t).delete()
            User.objects.filter(tenant=t).delete()
            t.refresh_from_db()
            t.delete()

        logger.info(
            "Superadmin %s deleted tenant %d (%s) with %d users",
            request.user.email, tenant_id, tenant_name, user_count,
        )
        return Response({"ok": True, "deleted": tenant_name, "users_deleted": user_count})


# ─────────────────────────────────────────────────────────────
# SUBSCRIPTION MANAGEMENT
# ─────────────────────────────────────────────────────────────
class AdminSubscriptionView(APIView):
    """Cambiar plan o estado de suscripción de un tenant. POST para crear si no existe."""
    permission_classes = PERMS

    def post(self, request, tenant_id):
        """Crear suscripción para tenant que no la tiene."""
        tenant = get_object_or_404(Tenant, pk=tenant_id)
        if Subscription.objects.filter(tenant=tenant).exists():
            return Response({"detail": "Este tenant ya tiene suscripción."}, status=400)

        plan_key = request.data.get("plan_key", "inicio")
        try:
            plan = Plan.objects.get(key=plan_key, is_active=True)
        except Plan.DoesNotExist:
            return Response({"detail": f"Plan '{plan_key}' no encontrado."}, status=400)

        status = request.data.get("status", "active")
        sub = Subscription.objects.create(
            tenant=tenant,
            plan=plan,
            status=status,
            current_period_start=timezone.now(),
            current_period_end=timezone.now() + timedelta(days=30),
        )
        logger.info("Superadmin %s created subscription for tenant %d: %s/%s",
                     request.user.email, tenant_id, plan_key, status)
        return Response({
            "ok": True, "status": sub.status, "plan": plan.name,
            "current_period_end": sub.current_period_end.isoformat(),
        }, status=201)

    def patch(self, request, tenant_id):
        sub = get_object_or_404(Subscription.objects.select_related("plan"), tenant_id=tenant_id)
        changed = []

        # Cambiar plan
        new_plan_key = request.data.get("plan_key")
        if new_plan_key and new_plan_key != sub.plan.key:
            try:
                new_plan = Plan.objects.get(key=new_plan_key, is_active=True)
                sub.plan = new_plan
                changed.append("plan")
                logger.info(
                    "Superadmin %s changed tenant %d plan to %s",
                    request.user.email, tenant_id, new_plan_key,
                )
            except Plan.DoesNotExist:
                return Response({"detail": f"Plan '{new_plan_key}' no encontrado."}, status=400)

        # Cambiar estado
        new_status = request.data.get("status")
        if new_status and new_status != sub.status:
            valid = [c[0] for c in Subscription.Status.choices]
            if new_status not in valid:
                return Response({"detail": f"Estado inválido: {new_status}"}, status=400)
            sub.status = new_status
            changed.append("status")
            if new_status == "suspended":
                sub.suspended_at = timezone.now()
                changed.append("suspended_at")
            elif new_status == "active":
                sub.suspended_at = None
                sub.payment_retry_count = 0
                changed.extend(["suspended_at", "payment_retry_count"])
            logger.info(
                "Superadmin %s changed tenant %d status to %s",
                request.user.email, tenant_id, new_status,
            )

        # Extender período
        extend_days = request.data.get("extend_days")
        if extend_days:
            days = safe_int(extend_days, 0)
            if sub.current_period_end:
                sub.current_period_end += timedelta(days=days)
            else:
                sub.current_period_end = timezone.now() + timedelta(days=days)
            changed.append("current_period_end")
            logger.info(
                "Superadmin %s extended tenant %d period by %d days",
                request.user.email, tenant_id, days,
            )

        if changed:
            sub.save(update_fields=list(set(changed)))

        return Response({
            "ok": True,
            "status": sub.status,
            "plan": sub.plan.key,
            "current_period_end": sub.current_period_end.isoformat() if sub.current_period_end else None,
        })


# ─────────────────────────────────────────────────────────────
# USER MANAGEMENT
# ─────────────────────────────────────────────────────────────
class AdminUserListView(APIView):
    """Lista todos los usuarios de la plataforma."""
    permission_classes = PERMS

    def get(self, request):
        qs = User.objects.filter(is_superuser=False).select_related("tenant").order_by("-date_joined")

        q = request.query_params.get("q", "").strip()
        if q:
            qs = qs.filter(
                Q(email__icontains=q) | Q(username__icontains=q) |
                Q(first_name__icontains=q) | Q(last_name__icontains=q)
            )

        role = request.query_params.get("role")
        if role:
            qs = qs.filter(role=role)

        active = request.query_params.get("active")
        if active == "true":
            qs = qs.filter(is_active=True)
        elif active == "false":
            qs = qs.filter(is_active=False)

        tenant_id = request.query_params.get("tenant_id")
        if tenant_id:
            qs = qs.filter(tenant_id=tenant_id)

        page = safe_int(request.query_params.get("page"), 1)
        page_size = safe_int(request.query_params.get("page_size"), 25)
        total = qs.count()
        offset = (page - 1) * page_size

        users = []
        for u in qs[offset:offset + page_size]:
            users.append({
                "id": u.id,
                "email": u.email,
                "username": u.username,
                "first_name": u.first_name,
                "last_name": u.last_name,
                "role": u.role,
                "is_active": u.is_active,
                "tenant_id": u.tenant_id,
                "tenant_name": u.tenant.name if u.tenant else None,
                "last_login": u.last_login.isoformat() if u.last_login else None,
                "date_joined": u.date_joined.isoformat(),
            })

        return Response({
            "results": users,
            "total": total,
            "page": page,
            "page_size": page_size,
        })


class AdminUserCreateView(APIView):
    """Crear usuario desde superadmin."""
    permission_classes = PERMS

    def post(self, request):
        email = request.data.get("email", "").strip()
        password = request.data.get("password", "").strip()
        first_name = request.data.get("first_name", "").strip()
        last_name = request.data.get("last_name", "").strip()
        role = request.data.get("role", "owner")
        tenant_id = request.data.get("tenant_id")

        if not email or not password:
            return Response({"detail": "Email y contraseña son requeridos."}, status=400)

        if not tenant_id:
            return Response({"detail": "Debes asignar un negocio (tenant) al usuario."}, status=400)

        if User.objects.filter(email=email).exists():
            return Response({"detail": "Ya existe un usuario con ese email."}, status=400)

        if User.objects.filter(username=email).exists():
            return Response({"detail": "Ya existe un usuario con ese nombre de usuario."}, status=400)

        tenant = get_object_or_404(Tenant, pk=tenant_id)

        try:
            user = User.objects.create_user(
                username=email,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                role=role,
                tenant=tenant,
            )
        except (IntegrityError, ValueError) as e:
            logger.warning("Error al crear usuario en superadmin: %s", e)
            return Response({"detail": "Error al crear el usuario. Verifica que los datos no estén duplicados."}, status=400)

        logger.info("Superadmin %s created user %s (tenant=%s)", request.user.email, email, tenant_id)

        return Response({
            "id": user.id,
            "email": user.email,
            "role": user.role,
            "tenant_id": user.tenant_id,
        }, status=201)


class AdminUserToggleView(APIView):
    """Activar/desactivar usuario."""
    permission_classes = PERMS

    def post(self, request, user_id):
        user = get_object_or_404(User, pk=user_id, is_superuser=False)
        user.is_active = not user.is_active
        user.save(update_fields=["is_active"])
        action = "activated" if user.is_active else "deactivated"
        logger.info("Superadmin %s %s user %d", request.user.email, action, user_id)
        return Response({"ok": True, "is_active": user.is_active})


class AdminUserDeleteView(APIView):
    """Eliminar usuario permanentemente, o desactivarlo si tiene datos asociados."""
    permission_classes = PERMS

    def delete(self, request, user_id):
        user = get_object_or_404(User, pk=user_id, is_superuser=False)
        username = user.username
        try:
            user.delete()
            logger.info("Superadmin %s deleted user %d (%s)", request.user.email, user_id, username)
            return Response({"ok": True, "deleted": username})
        except models.ProtectedError:
            user.is_active = False
            user.save(update_fields=["is_active"])
            logger.info("Superadmin %s deactivated user %d (%s) (has protected references)", request.user.email, user_id, username)
            return Response({"ok": True, "deactivated": username})


# ─────────────────────────────────────────────────────────────
# INVOICES (global view)
# ─────────────────────────────────────────────────────────────
class AdminInvoiceListView(APIView):
    permission_classes = PERMS

    def get(self, request):
        qs = Invoice.objects.select_related(
            "subscription__tenant", "subscription__plan"
        ).order_by("-created_at")

        inv_status = request.query_params.get("status")
        if inv_status:
            qs = qs.filter(status=inv_status)

        page = safe_int(request.query_params.get("page"), 1)
        page_size = safe_int(request.query_params.get("page_size"), 25)
        total = qs.count()
        offset = (page - 1) * page_size

        results = []
        for inv in qs[offset:offset + page_size]:
            results.append({
                "id": inv.id,
                "tenant_id": inv.subscription.tenant_id,
                "tenant_name": inv.subscription.tenant.name,
                "plan_name": inv.subscription.plan.name,
                "status": inv.status,
                "amount_clp": inv.amount_clp,
                "period_start": inv.period_start.isoformat(),
                "period_end": inv.period_end.isoformat(),
                "paid_at": inv.paid_at.isoformat() if inv.paid_at else None,
                "created_at": inv.created_at.isoformat(),
            })

        return Response({
            "results": results,
            "total": total,
            "page": page,
            "page_size": page_size,
        })


# ─────────────────────────────────────────────────────────────
# FORECAST METRICS (overview por tenant)
# ─────────────────────────────────────────────────────────────
class AdminForecastMetricsView(APIView):
    permission_classes = PERMS

    def get(self, request):
        try:
            from forecast.models import ForecastModel, ForecastAccuracy
        except ImportError:
            return Response({"detail": "Módulo forecast no disponible."}, status=404)

        mape_float = Cast(KeyTextTransform("mape", "metrics"), FloatField())

        active_models = ForecastModel.objects.filter(is_active=True)

        # Resumen global
        total_models = active_models.count()
        avg_mape = active_models.annotate(
            mape_val=mape_float,
        ).aggregate(avg=Avg("mape_val"))["avg"]

        # Por tenant (con avg MAPE)
        by_tenant = list(
            active_models.annotate(mape_val=mape_float).values(
                "tenant_id",
                tenant_name=F("tenant__name"),
            ).annotate(
                model_count=Count("id"),
                avg_mape=Avg("mape_val"),
            ).order_by("-model_count")[:50]
        )
        for t in by_tenant:
            t["avg_mape"] = round(float(t["avg_mape"] or 0), 2)

        # Distribución por algoritmo
        by_algorithm = list(
            active_models.values("algorithm").annotate(
                count=Count("id"),
            ).order_by("-count")
        )

        # Distribución por confianza
        by_confidence = list(
            active_models.values(confidence=F("confidence_label")).annotate(
                count=Count("id"),
            ).order_by("-count")
        )

        # Distribución por patrón de demanda
        by_pattern = list(
            active_models.values("demand_pattern").annotate(
                count=Count("id"),
            ).order_by("-count")
        )

        # Accuracy reciente (últimos 7 días)
        week_ago = timezone.now() - timedelta(days=7)
        recent_accuracy = ForecastAccuracy.objects.filter(
            date__gte=week_ago.date(),
        ).aggregate(
            avg_error=Avg("abs_pct_error"),
            total_predictions=Count("id"),
        )

        # Tendencia de accuracy diaria (últimos 30 días)
        month_ago = timezone.now() - timedelta(days=30)
        accuracy_trend = list(
            ForecastAccuracy.objects.filter(
                date__gte=month_ago.date(),
            ).values("date").annotate(
                avg_error=Avg("abs_pct_error"),
                predictions=Count("id"),
            ).order_by("date")
        )
        for row in accuracy_trend:
            row["date"] = str(row["date"])
            row["avg_error"] = round(float(row["avg_error"] or 0), 2)

        # Modelos mejorados vs empeorados (con mape_delta)
        improved = active_models.filter(mape_delta__lt=0).count()
        worsened = active_models.filter(mape_delta__gt=0).count()
        unchanged = active_models.filter(mape_delta=0).count()

        return Response({
            "total_active_models": total_models,
            "global_avg_mape": round(float(avg_mape or 0), 2),
            "by_tenant": by_tenant,
            "by_algorithm": by_algorithm,
            "by_confidence": by_confidence,
            "by_pattern": by_pattern,
            "model_health": {
                "improved": improved,
                "worsened": worsened,
                "unchanged": unchanged,
            },
            "accuracy_trend": accuracy_trend,
            "recent_7d": {
                "avg_pct_error": round(float(recent_accuracy["avg_error"] or 0), 2),
                "total_predictions": recent_accuracy["total_predictions"],
            },
            "training_logs": self._get_training_logs(),
        })

    def _get_training_logs(self):
        """Últimas 10 ejecuciones del pipeline de forecast."""
        from forecast.models import ForecastTrainingLog
        logs = ForecastTrainingLog.objects.order_by("-started_at")[:10]
        return [
            {
                "command": log.command,
                "status": log.status,
                "started_at": log.started_at.isoformat(),
                "duration_seconds": log.duration_seconds,
                "models_trained": log.models_trained,
                "models_improved": log.models_improved,
                "models_failed": log.models_failed,
                "avg_mape": log.avg_mape,
                "error_message": log.error_message[:200] if log.error_message else "",
                "algorithm_distribution": log.algorithm_distribution,
            }
            for log in logs
        ]


# ─────────────────────────────────────────────────────────────
# HOLIDAYS MANAGEMENT (Global)
# ─────────────────────────────────────────────────────────────
class AdminHolidayListView(APIView):
    """CRUD de holidays globales (tenant=null) y por tenant."""
    permission_classes = PERMS

    def get(self, request):
        from forecast.models import Holiday
        qs = Holiday.objects.all().order_by("date")
        scope = request.query_params.get("scope")
        if scope == "national":
            qs = qs.filter(tenant__isnull=True)
        elif scope == "custom":
            qs = qs.filter(tenant__isnull=False)

        tenant_id = request.query_params.get("tenant_id")
        if tenant_id:
            qs = qs.filter(Q(tenant_id=tenant_id) | Q(tenant__isnull=True))

        return Response([
            {
                "id": h.id,
                "name": h.name,
                "date": str(h.date),
                "scope": h.scope,
                "tenant_id": h.tenant_id,
                "demand_multiplier": str(h.demand_multiplier),
                "pre_days": h.pre_days,
                "pre_multiplier": str(h.pre_multiplier),
                "duration_days": h.duration_days,
                "post_days": h.post_days,
                "post_multiplier": str(h.post_multiplier),
                "ramp_type": h.ramp_type,
                "is_recurring": h.is_recurring,
                "learned_multiplier": str(h.learned_multiplier) if h.learned_multiplier else None,
            }
            for h in qs
        ])

    def post(self, request):
        from forecast.models import Holiday
        from decimal import Decimal
        data = request.data
        name = (data.get("name") or "").strip()
        date_str = data.get("date", "")
        if not name or not date_str:
            return Response({"detail": "Nombre y fecha son requeridos."}, status=400)

        from datetime import date as date_cls
        try:
            h_date = date_cls.fromisoformat(date_str)
        except ValueError:
            return Response({"detail": "Fecha inválida. Formato: YYYY-MM-DD"}, status=400)

        h = Holiday.objects.create(
            tenant_id=data.get("tenant_id"),  # null = national
            name=name,
            date=h_date,
            scope=data.get("scope", "NATIONAL"),
            demand_multiplier=Decimal(str(data.get("demand_multiplier", "1.50"))),
            pre_days=int(data.get("pre_days", 1)),
            pre_multiplier=Decimal(str(data.get("pre_multiplier", "1.20"))),
            duration_days=int(data.get("duration_days", 1)),
            post_days=int(data.get("post_days", 0)),
            post_multiplier=Decimal(str(data.get("post_multiplier", "0.85"))),
            ramp_type=data.get("ramp_type", "instant"),
            is_recurring=data.get("is_recurring", True),
        )
        return Response({"ok": True, "id": h.id, "name": h.name}, status=201)


class AdminHolidayDetailView(APIView):
    permission_classes = PERMS

    def patch(self, request, holiday_id):
        from forecast.models import Holiday
        from decimal import Decimal
        h = get_object_or_404(Holiday, pk=holiday_id)
        changed = []
        for field in ["name", "scope", "ramp_type"]:
            if field in request.data:
                setattr(h, field, request.data[field])
                changed.append(field)
        for field in ["pre_days", "duration_days", "post_days"]:
            if field in request.data:
                setattr(h, field, int(request.data[field]))
                changed.append(field)
        for field in ["demand_multiplier", "pre_multiplier", "post_multiplier"]:
            if field in request.data:
                setattr(h, field, Decimal(str(request.data[field])))
                changed.append(field)
        if "date" in request.data:
            from datetime import date as date_cls
            h.date = date_cls.fromisoformat(request.data["date"])
            changed.append("date")
        if "is_recurring" in request.data:
            h.is_recurring = bool(request.data["is_recurring"])
            changed.append("is_recurring")
        if changed:
            h.save(update_fields=changed)
        return Response({"ok": True, "id": h.id})

    def delete(self, request, holiday_id):
        from forecast.models import Holiday
        h = get_object_or_404(Holiday, pk=holiday_id)
        h.delete()
        return Response({"ok": True})


# ─────────────────────────────────────────────────────────────
# INVOICE MANUAL PAYMENT
# ─────────────────────────────────────────────────────────────
class AdminInvoicePayView(APIView):
    """Marcar factura como pagada manualmente."""
    permission_classes = PERMS

    def post(self, request, invoice_id):
        invoice = get_object_or_404(Invoice, pk=invoice_id)
        if invoice.status == "paid":
            return Response({"detail": "Esta factura ya está pagada."}, status=400)

        invoice.status = "paid"
        invoice.paid_at = timezone.now()
        invoice.gateway = "manual"
        invoice.save(update_fields=["status", "paid_at", "gateway"])

        logger.info("Superadmin %s marked invoice %d as paid (manual)",
                     request.user.email, invoice_id)
        return Response({"ok": True, "id": invoice.id, "status": "paid"})


# ─────────────────────────────────────────────────────────────
# FORECAST TRAINING (manual trigger)
# ─────────────────────────────────────────────────────────────
class AdminForecastTrainView(APIView):
    """POST /superadmin/forecast/train/ — trigger training from UI.
    Body: { tenant_id?: int, product_id?: int }
    """
    permission_classes = PERMS

    def post(self, request):
        from django.core.management import call_command
        from io import StringIO
        import time

        tenant_id = request.data.get("tenant_id")
        product_id = request.data.get("product_id")

        args = []
        if tenant_id:
            args += ["--tenant", str(tenant_id)]
        if product_id:
            args += ["--product", str(product_id)]

        scope = f"tenant={tenant_id}" if tenant_id else "todos los tenants"
        if product_id:
            scope += f", product={product_id}"

        logger.info("Superadmin %s triggered forecast training: %s", request.user.email, scope)

        start = time.time()
        try:
            out = StringIO()
            call_command("train_forecast_models", *args, stdout=out)
            elapsed = round(time.time() - start, 1)
            output = out.getvalue()
            logger.info("Training completed in %.1fs: %s", elapsed, output[-300:])

            # Parse output for summary numbers
            lines = output.strip().split("\n")
            summary = lines[-1] if lines else ""

            return Response({
                "ok": True,
                "message": f"Entrenamiento completado en {elapsed}s ({scope}).",
                "detail": summary,
                "elapsed": elapsed,
            })
        except Exception as e:
            elapsed = round(time.time() - start, 1)
            logger.error("Training failed after %.1fs: %s", elapsed, e)
            return Response({
                "ok": False,
                "message": f"Error en entrenamiento ({elapsed}s): {str(e)[:200]}",
            }, status=500)


# ─────────────────────────────────────────────────────────────
# IMPORT HISTORICAL SALES (CSV)
# ─────────────────────────────────────────────────────────────
class AdminImportSalesView(APIView):
    """POST /superadmin/forecast/import-sales/
    Multipart form: tenant_id + file (CSV/Excel)
    CSV columns: date, product_id (or sku or name), qty_sold, total (optional), promo_qty (optional)
    """
    permission_classes = PERMS

    # Cap de filas para prevenir DoS por archivo gigante (sin esto, un
    # CSV de 5MB con líneas cortas podía tener >200k filas y procesarlas
    # todas sin transacción → corruption parcial si reventaba a mitad).
    MAX_ROWS = 50000  # superadmin tiene casos legítimos con muchos meses
    MAX_FILE_SIZE = 20 * 1024 * 1024  # 20 MB
    MAX_ERRORS_RETURN = 100

    def post(self, request):
        import csv
        from datetime import date as date_cls
        from decimal import Decimal, InvalidOperation
        from django.db import transaction

        tenant_id = request.data.get("tenant_id")
        if not tenant_id:
            return Response({"detail": "tenant_id es requerido."}, status=400)

        try:
            tenant_id = int(tenant_id)
        except (ValueError, TypeError):
            return Response({"detail": "tenant_id inválido."}, status=400)

        from core.models import Tenant
        tenant = Tenant.objects.filter(id=tenant_id).first()
        if not tenant:
            return Response({"detail": "Tenant no encontrado."}, status=404)

        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response({"detail": "Archivo es requerido."}, status=400)

        # Defensa por tamaño antes de leer en memoria
        if uploaded.size > self.MAX_FILE_SIZE:
            mb = self.MAX_FILE_SIZE // (1024 * 1024)
            return Response({"detail": f"Archivo demasiado grande (máximo {mb} MB)."}, status=413)

        # Parse file
        fname = uploaded.name.lower()
        rows = []

        if fname.endswith(".csv"):
            import io
            # `errors="replace"` antes silenciosamente ponía `?` en
            # caracteres mal codificados → basura persistida. Mejor
            # fallar explícito para que el dueño re-exporte en UTF-8.
            try:
                content = uploaded.read().decode("utf-8-sig")
            except UnicodeDecodeError as e:
                return Response({
                    "detail": (
                        f"Encoding inválido: {e}. Guardá el CSV como UTF-8 "
                        f"(en Excel: Archivo → Guardar como → CSV UTF-8)."
                    )
                }, status=400)
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        elif fname.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            wb.close()
        else:
            return Response({"detail": "Formato no soportado. Usa CSV o Excel (.xlsx)."}, status=400)

        if not rows:
            return Response({"detail": "Archivo vacío."}, status=400)

        if len(rows) > self.MAX_ROWS:
            return Response({
                "detail": (
                    f"Archivo tiene {len(rows)} filas, máximo {self.MAX_ROWS}. "
                    f"Dividilo en archivos más chicos."
                )
            }, status=413)

        # Header mapping (Spanish support)
        HEADER_MAP = {
            "fecha": "date", "producto": "product_id", "producto_id": "product_id",
            "sku": "sku", "nombre": "name", "nombre_producto": "name",
            "cantidad": "qty_sold", "qty": "qty_sold", "qty_sold": "qty_sold",
            "venta": "total", "total": "total", "monto": "total",
            "promo": "promo_qty", "promo_qty": "promo_qty", "qty_promo": "promo_qty",
        }

        def norm_key(k):
            return HEADER_MAP.get(k.strip().lower().replace(" ", "_"), k.strip().lower())

        rows = [{norm_key(k): v for k, v in r.items()} for r in rows]

        # Resolve products
        from catalog.models import Product
        products_by_id = {p.id: p for p in Product.objects.filter(tenant_id=tenant_id)}
        products_by_sku = {p.sku.upper(): p for p in products_by_id.values() if p.sku}
        products_by_name = {p.name.upper(): p for p in products_by_id.values()}

        # Get or create warehouse
        from core.models import Warehouse
        warehouse = Warehouse.objects.filter(tenant_id=tenant_id, is_active=True).first()
        if not warehouse:
            return Response({"detail": "Tenant no tiene bodega activa."}, status=400)

        from forecast.models import DailySales

        created = 0
        updated = 0
        skipped = 0
        errors = []

        # Transacción global con savepoints por fila: si una fila revienta
        # se rollback solo esa, las anteriores quedan persistidas. Si falla
        # algo POST-loop (commit, audit log), revierte todo. Antes: el
        # bucle no estaba envuelto → un crash a mitad dejaba 50% importado
        # con corruption silenciosa.
        with transaction.atomic():
            for i, row in enumerate(rows, start=2):
                sid = transaction.savepoint()
                try:
                    # Parse date
                    dt_raw = row.get("date", "")
                    try:
                        if hasattr(dt_raw, "date") and not isinstance(dt_raw, date_cls):
                            dt = dt_raw.date()
                        elif isinstance(dt_raw, date_cls):
                            dt = dt_raw
                        else:
                            dt = date_cls.fromisoformat(str(dt_raw).strip()[:10])
                    except (ValueError, TypeError):
                        errors.append(f"Fila {i}: fecha inválida '{dt_raw}'")
                        skipped += 1
                        transaction.savepoint_rollback(sid)
                        continue

                    # Resolve product
                    product = None
                    pid = row.get("product_id")
                    sku = row.get("sku")
                    name = row.get("name")

                    if pid:
                        try:
                            product = products_by_id.get(int(pid))
                        except (ValueError, TypeError):
                            pass
                    if not product and sku:
                        product = products_by_sku.get(str(sku).strip().upper())
                    if not product and name:
                        product = products_by_name.get(str(name).strip().upper())

                    if not product:
                        errors.append(f"Fila {i}: producto no encontrado (id={pid}, sku={sku}, name={name})")
                        skipped += 1
                        transaction.savepoint_rollback(sid)
                        continue

                    # Parse qty — NO permitir negativos ni NaN/Inf
                    try:
                        qty_sold = Decimal(str(row.get("qty_sold", 0) or 0))
                        if not qty_sold.is_finite():
                            raise InvalidOperation("non-finite")
                    except (InvalidOperation, TypeError):
                        errors.append(f"Fila {i}: cantidad inválida")
                        skipped += 1
                        transaction.savepoint_rollback(sid)
                        continue
                    if qty_sold < 0:
                        errors.append(f"Fila {i}: qty_sold negativo no permitido ({qty_sold})")
                        skipped += 1
                        transaction.savepoint_rollback(sid)
                        continue

                    promo_qty = Decimal("0")
                    try:
                        promo_qty = Decimal(str(row.get("promo_qty", 0) or 0))
                        if not promo_qty.is_finite() or promo_qty < 0:
                            promo_qty = Decimal("0")
                    except (InvalidOperation, TypeError):
                        promo_qty = Decimal("0")

                    # Total: si viene mal, REPORTAR error (no fallback al
                    # precio actual del producto, que distorsiona revenue
                    # histórico). Si falta del todo, también error explícito.
                    total_raw = row.get("total")
                    if total_raw is None or str(total_raw).strip() == "":
                        # Total ausente → asumir 0 con warning explícito.
                        # No usamos product.price porque el precio de hoy
                        # no es el de hace 6 meses (inflación, promos).
                        total = Decimal("0")
                    else:
                        try:
                            total = Decimal(str(total_raw))
                            if not total.is_finite() or total < 0:
                                raise InvalidOperation("invalid")
                        except (InvalidOperation, TypeError):
                            errors.append(
                                f"Fila {i}: total inválido '{total_raw}'. "
                                f"Dejá la celda vacía si no tenés el dato (asumirá 0)."
                            )
                            skipped += 1
                            transaction.savepoint_rollback(sid)
                            continue

                    # Upsert DailySales
                    obj, was_created = DailySales.objects.update_or_create(
                        tenant_id=tenant_id,
                        product=product,
                        warehouse=warehouse,
                        date=dt,
                        defaults={
                            "qty_sold": qty_sold,
                            "revenue": total,
                            "promo_qty": promo_qty,
                        },
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                    transaction.savepoint_commit(sid)
                except Exception as e:
                    transaction.savepoint_rollback(sid)
                    errors.append(f"Fila {i}: {str(e)[:200]}")
                    skipped += 1

        logger.info(
            "Superadmin %s imported sales for tenant %d: %d created, %d updated, %d skipped",
            request.user.email, tenant_id, created, updated, skipped,
        )

        return Response({
            "ok": True,
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors[: self.MAX_ERRORS_RETURN],
            "errors_count": len(errors),
            "total_rows": len(rows),
        })


# ═════════════════════════════════════════════════════════════════════════════
# SOPORTE — Features para que el admin pueda asistir clientes en tiempo real
# (Fase: pre-piloto cafetería)
# ═════════════════════════════════════════════════════════════════════════════

class AdminUserResetPasswordView(APIView):
    """POST /api/superadmin/users/<id>/reset-password/

    Genera una contraseña aleatoria nueva para el user y la setea con
    set_password(). Devuelve la pass UNA sola vez en la respuesta para que
    el admin se la dicte/envíe al cliente. Después de este request, la pass
    nueva ya no puede recuperarse — solo el cliente la conoce (y el admin
    si la guardó del response).

    Pensado para el escenario "cliente olvidó su contraseña" sin tener un
    flujo automático de password reset por email todavía.
    """
    permission_classes = PERMS

    def post(self, request, user_id):
        import secrets, string
        user = get_object_or_404(User, pk=user_id)

        # Pass legible: 12 chars, sin caracteres confusos (0/O, 1/l/I)
        alphabet = string.ascii_letters + string.digits
        alphabet = "".join(c for c in alphabet if c not in "0OIl1")
        new_password = "".join(secrets.choice(alphabet) for _ in range(12))

        user.set_password(new_password)
        user.save(update_fields=["password"])

        logger.info(
            "Superadmin %s reset password for user %d (%s)",
            request.user.email, user.pk, user.email,
        )
        return Response({
            "ok": True,
            "user_id": user.pk,
            "email": user.email,
            "new_password": new_password,
            "warning": "Esta contraseña solo se muestra UNA VEZ. Anótela ahora.",
        })


class AdminTenantResendEmailView(APIView):
    """POST /api/superadmin/tenants/<id>/resend-email/

    Body: { "type": "welcome" | "payment_link" | "trial_reminder" |
                     "payment_recovered" | "renewal_reminder" }

    Reenvía un email transaccional al owner del tenant. Útil cuando el
    cliente reporta "no me llegó". Usa los renderers existentes para
    asegurar consistencia con los emails automáticos.
    """
    permission_classes = PERMS

    EMAIL_TYPES = {
        "welcome", "payment_link", "trial_reminder",
        "payment_recovered", "renewal_reminder",
    }

    def post(self, request, tenant_id):
        email_type = (request.data.get("type") or "").strip().lower()
        if email_type not in self.EMAIL_TYPES:
            return Response(
                {"detail": f"type debe ser uno de: {', '.join(sorted(self.EMAIL_TYPES))}"},
                status=400,
            )

        tenant = get_object_or_404(Tenant, pk=tenant_id)

        owner = User.objects.filter(
            tenant=tenant, role="owner", is_active=True,
        ).first()
        if not owner or not owner.email:
            return Response(
                {"detail": "El tenant no tiene un owner activo con email."},
                status=400,
            )

        sub = Subscription.objects.filter(tenant=tenant).select_related("plan").first()

        try:
            from billing import tasks as T
            from django.utils import timezone as _tz

            if email_type == "welcome":
                if not sub:
                    return Response({"detail": "El tenant no tiene suscripción."}, status=400)
                T._send_welcome_email(owner, tenant, sub.plan)
            elif email_type == "trial_reminder":
                if not sub or not sub.trial_ends_at:
                    return Response({"detail": "El tenant no tiene un trial activo."}, status=400)
                days_left = max(0, (sub.trial_ends_at - _tz.now()).days)
                T._send_trial_reminder(sub, days_left)
            elif email_type == "renewal_reminder":
                if not sub:
                    return Response({"detail": "El tenant no tiene suscripción."}, status=400)
                days_left = max(
                    0,
                    (sub.current_period_end - _tz.now()).days if sub.current_period_end else 0,
                )
                T._send_renewal_reminder(sub, days_left)
            elif email_type == "payment_recovered":
                if not sub:
                    return Response({"detail": "El tenant no tiene suscripción."}, status=400)
                T._send_payment_recovered_notice(sub)
            elif email_type == "payment_link":
                if not sub:
                    return Response({"detail": "El tenant no tiene suscripción."}, status=400)
                T._send_payment_failed_notice(sub)
        except Exception as e:
            logger.exception("Resend email %s failed for tenant %d", email_type, tenant_id)
            return Response(
                {"detail": f"El email no pudo enviarse: {e}"},
                status=502,
            )

        logger.info(
            "Superadmin %s resent email '%s' to %s (tenant=%d)",
            request.user.email, email_type, owner.email, tenant_id,
        )
        return Response({
            "ok": True,
            "type": email_type,
            "sent_to": owner.email,
        })


class AdminGlobalSearchView(APIView):
    """GET /api/superadmin/search/?q=texto

    Búsqueda global rápida para soporte. Busca en:
      - Tenant: name, slug, rut, email, legal_name
      - User: username, email, first_name, last_name

    Devuelve hasta 20 resultados de cada tipo. q debe tener al menos 2
    caracteres para evitar listados accidentales.
    """
    permission_classes = PERMS

    def get(self, request):
        q = (request.query_params.get("q") or "").strip()
        if len(q) < 2:
            return Response({"tenants": [], "users": [], "query": q})

        tenants_qs = Tenant.objects.filter(
            Q(name__icontains=q)
            | Q(slug__icontains=q)
            | Q(rut__icontains=q)
            | Q(email__icontains=q)
            | Q(legal_name__icontains=q)
        ).order_by("-id")[:20]

        users_qs = User.objects.filter(
            Q(username__icontains=q)
            | Q(email__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
        ).select_related("tenant").order_by("-id")[:20]

        return Response({
            "query": q,
            "tenants": [
                {
                    "id": t.id, "name": t.name, "slug": t.slug,
                    "rut": t.rut, "email": t.email, "is_active": t.is_active,
                }
                for t in tenants_qs
            ],
            "users": [
                {
                    "id": u.id, "username": u.username, "email": u.email,
                    "full_name": (f"{u.first_name} {u.last_name}").strip(),
                    "role": u.role, "is_active": u.is_active,
                    "tenant": {"id": u.tenant_id, "name": u.tenant.name} if u.tenant else None,
                }
                for u in users_qs
            ],
        })


class AdminTenantNotesView(APIView):
    """PATCH /api/superadmin/tenants/<id>/notes/

    Actualiza el campo internal_notes del tenant. Soporte interno (no
    visible al cliente). Útil para el admin para anotar contexto del
    cliente: "llamó por X, prefiere WhatsApp, etc."

    Body: { "internal_notes": "texto..." }
    """
    permission_classes = PERMS

    def patch(self, request, tenant_id):
        tenant = get_object_or_404(Tenant, pk=tenant_id)
        notes = request.data.get("internal_notes", "")
        if not isinstance(notes, str):
            return Response({"detail": "internal_notes debe ser string."}, status=400)

        if len(notes) > 5000:
            return Response(
                {"detail": "Las notas no pueden exceder 5000 caracteres."},
                status=400,
            )

        tenant.internal_notes = notes
        tenant.save(update_fields=["internal_notes"])

        logger.info(
            "Superadmin %s updated internal_notes for tenant %d (%d chars)",
            request.user.email, tenant.id, len(notes),
        )
        return Response({
            "ok": True,
            "tenant_id": tenant.id,
            "internal_notes": tenant.internal_notes,
        })
