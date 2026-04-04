"""
Report export views — Excel + PDF downloads for each report.
Uses openpyxl for .xlsx and reportlab for .pdf.
"""
from io import BytesIO
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from core.permissions import HasTenant, IsManager

from reports import services

import logging
logger = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center")


def _make_wb(title, headers, rows, filename):
    """Create an Excel workbook with styled headers and auto-width columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _require_ctx(request):
    t_id = getattr(request.user, "tenant_id", None)
    s_id = getattr(request.user, "active_store_id", None)
    if not t_id or not s_id:
        from rest_framework.exceptions import ValidationError
        raise ValidationError("Contexto de tienda no disponible.")
    return t_id, s_id


def _parse_dates(request):
    now = timezone.now().date()
    from datetime import timedelta
    date_from = parse_date(request.query_params.get("date_from", "")) or (now - timedelta(days=30))
    date_to = parse_date(request.query_params.get("date_to", "")) or now
    return date_from, date_to


class SalesSummaryExportView(APIView):
    permission_classes = [IsAuthenticated, HasTenant, IsManager]

    def get(self, request):
        t_id, s_id = _require_ctx(request)
        date_from, date_to = _parse_dates(request)
        data = services.get_sales_summary(t_id, s_id, date_from=date_from, date_to=date_to)

        headers = ["Fecha", "Ventas", "Ingresos", "Costo", "Ganancia", "Margen %"]
        rows = []
        for day in data.get("daily", []):
            revenue = float(day.get("total", 0))
            cost = float(day.get("cost", 0))
            profit = float(day.get("profit", 0))
            margin = round((profit / revenue * 100), 1) if revenue else 0
            rows.append([
                day.get("date", ""),
                day.get("count", 0),
                revenue,
                cost,
                profit,
                margin,
            ])

        return _make_wb("Resumen Ventas", headers, rows, f"ventas_{date_from}_{date_to}.xlsx")


class LossesExportView(APIView):
    permission_classes = [IsAuthenticated, HasTenant, IsManager]

    def get(self, request):
        t_id, s_id = _require_ctx(request)
        date_from, date_to = _parse_dates(request)
        wh_id = request.query_params.get("warehouse_id")
        reason = request.query_params.get("reason")

        data = services.get_losses(
            t_id, s_id,
            date_from=date_from, date_to=date_to,
            warehouse_id=int(wh_id) if wh_id and wh_id.isdigit() else None,
            reason=reason,
        )

        headers = ["Fecha", "Producto", "SKU", "Bodega", "Razón", "Cantidad", "Costo perdido"]
        rows = []
        for item in data.get("details", []):
            rows.append([
                item.get("date", ""),
                item.get("product_name", ""),
                item.get("sku", ""),
                item.get("warehouse_name", ""),
                item.get("reason", ""),
                float(item.get("qty", 0)),
                float(item.get("cost", 0)),
            ])

        return _make_wb("Mermas", headers, rows, f"mermas_{date_from}_{date_to}.xlsx")


class StockValuedExportView(APIView):
    permission_classes = [IsAuthenticated, HasTenant, IsManager]

    def get(self, request):
        t_id, s_id = _require_ctx(request)
        wh_id = request.query_params.get("warehouse_id")
        q = (request.query_params.get("q") or "").strip() or None

        data = services.get_stock_valued(
            t_id, s_id,
            warehouse_id=int(wh_id) if wh_id and wh_id.isdigit() else None,
            q=q,
        )

        headers = ["Bodega", "SKU", "Producto", "Stock", "Costo promedio", "Valor stock"]
        rows = []
        for wh in data.get("warehouses", []):
            wh_name = wh.get("warehouse_name", "")
            for item in wh.get("items", []):
                rows.append([
                    wh_name,
                    item.get("sku", ""),
                    item.get("product_name", ""),
                    float(item.get("on_hand", 0)),
                    float(item.get("avg_cost", 0)),
                    float(item.get("stock_value", 0)),
                ])

        return _make_wb("Stock Valorizado", headers, rows, "stock_valorizado.xlsx")


class AuditTrailExportView(APIView):
    permission_classes = [IsAuthenticated, HasTenant, IsManager]

    def get(self, request):
        t_id, s_id = _require_ctx(request)
        date_from, date_to = _parse_dates(request)
        ref_type = request.query_params.get("ref_type")
        move_type = request.query_params.get("move_type")

        from inventory.models import StockMove
        from django.db.models import Q

        qs = StockMove.objects.filter(
            tenant_id=t_id,
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
        ).select_related("product", "warehouse", "created_by").order_by("-created_at")

        if ref_type:
            qs = qs.filter(ref_type=ref_type)
        if move_type:
            qs = qs.filter(move_type=move_type)

        # Cap at 5000 for export
        qs = qs[:5000]

        headers = ["Fecha", "Tipo", "Referencia", "Producto", "Bodega", "Cantidad", "Valor", "Usuario", "Nota"]
        rows = []
        for m in qs:
            rows.append([
                m.created_at.strftime("%Y-%m-%d %H:%M") if m.created_at else "",
                m.move_type,
                m.ref_type,
                m.product.name if m.product else "",
                m.warehouse.name if m.warehouse else "",
                float(m.qty or 0),
                float(m.value_delta or 0),
                (m.created_by.get_full_name() or m.created_by.email) if m.created_by else "",
                m.note or "",
            ])

        return _make_wb("Auditoría", headers, rows, f"auditoria_{date_from}_{date_to}.xlsx")


# ══════════════════════════════════════════════════════════════════════════
# PDF EXPORTS (reportlab)
# ══════════════════════════════════════════════════════════════════════════

def _make_pdf(title, headers, rows, filename, landscape_mode=False):
    """Create a styled PDF table using reportlab. Returns HttpResponse."""
    from reportlab.lib.pagesizes import letter, landscape as lr_landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = BytesIO()
    page = lr_landscape(letter) if landscape_mode else letter
    doc = SimpleDocTemplate(buf, pagesize=page,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Heading1"],
        fontSize=16, textColor=colors.HexColor("#4F46E5"),
        spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#71717A"),
        spaceAfter=14,
    )
    cell_style = ParagraphStyle(
        "CellText", parent=styles["Normal"], fontSize=8, leading=10,
    )

    elements = []
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(
        f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')} — Pulstock",
        subtitle_style,
    ))

    # Truncate long cell values and wrap in Paragraphs
    def _cell(val):
        s = str(val) if val is not None else ""
        if len(s) > 40:
            s = s[:37] + "..."
        return Paragraph(s, cell_style)

    table_data = [[Paragraph(str(h), cell_style) for h in headers]]
    for row in rows[:2000]:  # cap at 2000 rows
        table_data.append([_cell(v) for v in row])

    n_cols = len(headers)
    avail_width = page[0] - 30*mm
    col_width = avail_width / n_cols

    table = Table(table_data, colWidths=[col_width] * n_cols, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F8")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4E4E7")),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)

    response = HttpResponse(buf.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


class SalesSummaryPDFView(APIView):
    permission_classes = [IsAuthenticated, HasTenant, IsManager]

    def get(self, request):
        t_id, s_id = _require_ctx(request)
        date_from, date_to = _parse_dates(request)
        data = services.get_sales_summary(t_id, s_id, date_from=date_from, date_to=date_to)

        headers = ["Fecha", "Ventas", "Ingresos", "Costo", "Ganancia", "Margen %"]
        rows = []
        for day in data.get("daily", []):
            revenue = float(day.get("total", 0))
            cost = float(day.get("cost", 0))
            profit = float(day.get("profit", 0))
            margin = round((profit / revenue * 100), 1) if revenue else 0
            rows.append([day.get("date", ""), day.get("count", 0),
                         f"${revenue:,.0f}", f"${cost:,.0f}", f"${profit:,.0f}", f"{margin}%"])

        return _make_pdf("Resumen de Ventas", headers, rows,
                         f"ventas_{date_from}_{date_to}.pdf")


class StockValuedPDFView(APIView):
    permission_classes = [IsAuthenticated, HasTenant, IsManager]

    def get(self, request):
        t_id, s_id = _require_ctx(request)
        wh_id = request.query_params.get("warehouse_id")
        data = services.get_stock_valued(
            t_id, s_id,
            warehouse_id=int(wh_id) if wh_id and wh_id.isdigit() else None,
        )

        headers = ["Bodega", "SKU", "Producto", "Stock", "Costo prom.", "Valor"]
        rows = []
        for wh in data.get("warehouses", []):
            wh_name = wh.get("warehouse_name", "")
            for item in wh.get("items", []):
                rows.append([wh_name, item.get("sku", ""), item.get("product_name", ""),
                             float(item.get("on_hand", 0)),
                             f"${float(item.get('avg_cost', 0)):,.0f}",
                             f"${float(item.get('stock_value', 0)):,.0f}"])

        return _make_pdf("Stock Valorizado", headers, rows,
                         "stock_valorizado.pdf", landscape_mode=True)


class LossesPDFView(APIView):
    permission_classes = [IsAuthenticated, HasTenant, IsManager]

    def get(self, request):
        t_id, s_id = _require_ctx(request)
        date_from, date_to = _parse_dates(request)
        wh_id = request.query_params.get("warehouse_id")
        reason = request.query_params.get("reason")

        data = services.get_losses(
            t_id, s_id, date_from=date_from, date_to=date_to,
            warehouse_id=int(wh_id) if wh_id and wh_id.isdigit() else None,
            reason=reason,
        )

        headers = ["Fecha", "Producto", "SKU", "Bodega", "Razón", "Cant.", "Costo"]
        rows = []
        for item in data.get("details", []):
            rows.append([item.get("date", ""), item.get("product_name", ""),
                         item.get("sku", ""), item.get("warehouse_name", ""),
                         item.get("reason", ""), float(item.get("qty", 0)),
                         f"${float(item.get('cost', 0)):,.0f}"])

        return _make_pdf("Mermas / Pérdidas", headers, rows,
                         f"mermas_{date_from}_{date_to}.pdf")


class AuditTrailPDFView(APIView):
    permission_classes = [IsAuthenticated, HasTenant, IsManager]

    def get(self, request):
        t_id, s_id = _require_ctx(request)
        date_from, date_to = _parse_dates(request)
        ref_type = request.query_params.get("ref_type")
        move_type = request.query_params.get("move_type")

        from inventory.models import StockMove

        qs = StockMove.objects.filter(
            tenant_id=t_id,
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
        ).select_related("product", "warehouse", "created_by").order_by("-created_at")

        if ref_type:
            qs = qs.filter(ref_type=ref_type)
        if move_type:
            qs = qs.filter(move_type=move_type)
        qs = qs[:2000]

        headers = ["Fecha", "Tipo", "Ref", "Producto", "Bodega", "Cant.", "Valor", "Usuario", "Nota"]
        rows = []
        for m in qs:
            rows.append([
                m.created_at.strftime("%Y-%m-%d %H:%M") if m.created_at else "",
                m.move_type, m.ref_type,
                m.product.name if m.product else "",
                m.warehouse.name if m.warehouse else "",
                float(m.qty or 0), f"${float(m.value_delta or 0):,.0f}",
                (m.created_by.get_full_name() or m.created_by.email) if m.created_by else "",
                (m.note or "")[:30],
            ])

        return _make_pdf("Auditoría de Movimientos", headers, rows,
                         f"auditoria_{date_from}_{date_to}.pdf", landscape_mode=True)


class ABCAnalysisPDFView(APIView):
    permission_classes = [IsAuthenticated, HasTenant, IsManager]

    def get(self, request):
        t_id, s_id = _require_ctx(request)
        date_from, date_to = _parse_dates(request)
        criterion = request.query_params.get("criterion", "revenue")

        data = services.get_abc_analysis(
            t_id, s_id, criterion=criterion,
            date_from=date_from, date_to=date_to,
        )

        headers = ["#", "Producto", "SKU", "Clase", "Revenue", "Costo", "Utilidad", "Margen", "Contrib.%"]
        rows = []
        for item in data.get("items", []):
            rows.append([
                item.get("rank", ""), item.get("product_name", ""),
                item.get("sku", ""), item.get("abc_class", ""),
                f"${float(item.get('revenue', 0)):,.0f}",
                f"${float(item.get('cost', 0)):,.0f}",
                f"${float(item.get('profit', 0)):,.0f}",
                f"{item.get('margin_pct', 0)}%",
                f"{item.get('contribution_pct', 0)}%",
            ])

        return _make_pdf("Análisis ABC", headers, rows,
                         f"abc_{criterion}_{date_from}_{date_to}.pdf", landscape_mode=True)
