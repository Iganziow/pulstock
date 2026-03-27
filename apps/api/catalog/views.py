from django.db.models import Q, Prefetch, Exists, OuterRef, Count, F, Case, When, Value
from django.db.models.functions import Greatest
from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from decimal import Decimal,InvalidOperation
from core.permissions import HasTenant, IsManager
from .models import Category, Product, Barcode, Recipe, RecipeLine, Unit
from .serializers import (
    CategorySerializer, ProductReadSerializer, ProductWriteSerializer,
    RecipeReadSerializer, RecipeWriteSerializer, UnitSerializer,
)
import csv
import io
from django.db import transaction


def tenant_id(request):
    return request.user.tenant_id


def _check_product_limit(t_id):
    """Verifica límite de productos del plan. Si no hay suscripción, permite."""
    try:
        from billing.models import Subscription
        from billing.services import check_plan_limit
        sub = Subscription.objects.select_related("plan").get(tenant_id=t_id)
        current = Product.objects.filter(tenant_id=t_id, is_active=True).count()
        return check_plan_limit(sub, "products", current)
    except Exception:
        return {"allowed": True, "limit": -1, "current": 0}


def _prefetch_barcodes_ordered():
    # orden estable (por code o id). Elige 1.
    return Prefetch("barcodes", queryset=Barcode.objects.order_by("code"))


# -----------------------
# Categories
# -----------------------
class CategoryListCreate(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated, HasTenant]
    serializer_class = CategorySerializer

    def get_queryset(self):
        return (
            Category.objects
            .filter(tenant_id=tenant_id(self.request))
            .select_related("parent")
            .annotate(children_count=Count("children", distinct=True))
            .order_by("name")
        )

    def perform_create(self, serializer):
        serializer.save(tenant_id=tenant_id(self.request))


class CategoryDetail(generics.RetrieveUpdateAPIView):
    permission_classes = [IsAuthenticated, HasTenant]
    serializer_class = CategorySerializer

    def get_queryset(self):
        return Category.objects.filter(tenant_id=tenant_id(self.request)).order_by("name")


# -----------------------
# Products
# -----------------------
class ProductListCreate(generics.ListCreateAPIView):
    """
    Catálogo:
    - GET /api/catalog/products/?q=
    - POST /api/catalog/products/
    """
    permission_classes = [IsAuthenticated, HasTenant]

    def get_queryset(self):
        qs = (
            Product.objects
            .filter(tenant_id=tenant_id(self.request))
            .select_related("category", "unit_obj", "recipe")
            .prefetch_related(_prefetch_barcodes_ordered())
        )

        q = (self.request.query_params.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(name__icontains=q) |
                Q(sku__icontains=q) |
                Exists(
                    Barcode.objects.filter(
                        product_id=OuterRef("pk"),
                        tenant_id=tenant_id(self.request),
                        code__icontains=q,
                    )
                )
            )

        return qs.order_by("name")

    def get_serializer_class(self):
        return ProductWriteSerializer if self.request.method == "POST" else ProductReadSerializer

    def perform_create(self, serializer):
        serializer.save(tenant_id=tenant_id(self.request))

    def create(self, request, *args, **kwargs):
        t_id = tenant_id(request)
        limit_check = _check_product_limit(t_id)
        if not limit_check["allowed"]:
            return Response(
                {
                    "detail": f"Has alcanzado el límite de {limit_check['limit']} productos de tu plan.",
                    "limit": limit_check["limit"],
                    "current": limit_check["current"],
                },
                status=status.HTTP_402_PAYMENT_REQUIRED,
            )
        write_ser = self.get_serializer(data=request.data)
        write_ser.is_valid(raise_exception=True)
        self.perform_create(write_ser)

        obj = (
            Product.objects
            .filter(pk=write_ser.instance.pk, tenant_id=tenant_id(self.request))
            .select_related("category")
            .prefetch_related(_prefetch_barcodes_ordered())
            .first()
        )

        read_ser = ProductReadSerializer(obj)
        headers = self.get_success_headers(read_ser.data)
        return Response(read_ser.data, status=status.HTTP_201_CREATED, headers=headers)


class ProductDetail(generics.RetrieveUpdateAPIView):
    permission_classes = [IsAuthenticated, HasTenant]

    def get_queryset(self):
        return (
            Product.objects
            .filter(tenant_id=tenant_id(self.request))
            .select_related("category")
            .prefetch_related(_prefetch_barcodes_ordered())
        )

    def get_serializer_class(self):
        return ProductWriteSerializer if self.request.method in ("PUT", "PATCH") else ProductReadSerializer

    # ✅ para que PATCH/PUT respondan con el read serializer (consistente con la tabla)
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()

        write_ser = ProductWriteSerializer(instance, data=request.data, partial=partial, context={"request": request})
        write_ser.is_valid(raise_exception=True)
        self.perform_update(write_ser)

        obj = (
            Product.objects
            .filter(pk=instance.pk, tenant_id=tenant_id(self.request))
            .select_related("category")
            .prefetch_related(_prefetch_barcodes_ordered())
            .first()
        )

        read_ser = ProductReadSerializer(obj)
        return Response(read_ser.data, status=status.HTTP_200_OK)


# -----------------------
# Search / Lookup (compat + POS)
# -----------------------
class ProductSearch(APIView):
    """
    Compat (no romper lo existente):
    - Si viene term: comportamiento POS (exact match) pero devuelve {"results":[...]}
    - Si viene q: búsqueda libre (top 10) devuelve {"results":[...]}

    Recomendación:
    - Catálogo use /products/?q=
    - POS use /products/lookup/?term=
    """
    permission_classes = [IsAuthenticated, HasTenant]

    def get(self, request):
        t_id = tenant_id(request)

        q = (request.query_params.get("q") or "").strip()
        term = (request.query_params.get("term") or "").strip()

        # 1) term => exact match (POS style)
        if term:
            bc = (
                Barcode.objects
                .filter(tenant_id=t_id, code=term)
                .select_related("product__category")
                .prefetch_related(Prefetch("product__barcodes", queryset=Barcode.objects.order_by("code")))
                .first()
            )
            if bc and bc.product:
                return Response({"results": [ProductReadSerializer(bc.product).data]})

            p = (
                Product.objects
                .filter(tenant_id=t_id, sku__iexact=term)
                .select_related("category")
                .prefetch_related(_prefetch_barcodes_ordered())
                .first()
            )
            if p:
                return Response({"results": [ProductReadSerializer(p).data]})

            return Response({"results": []})

        # 2) q => search libre
        if not q:
            return Response({"results": []})

        qs = (
            Product.objects
            .filter(tenant_id=t_id)
            .select_related("category")
            .prefetch_related(_prefetch_barcodes_ordered())
            .filter(
                Q(name__icontains=q) |
                Q(sku__icontains=q) |
                Q(barcodes__code__icontains=q)
            )
            .distinct()
            .order_by("name")[:10]
        )
        return Response({"results": ProductReadSerializer(qs, many=True).data})


class ProductLookup(APIView):
    """
    POS: lookup exacto por barcode o sku.
    Devuelve 1 producto (objeto) o 404.

    GET /api/catalog/products/lookup/?term=XXXXXXXX
    """
    permission_classes = [IsAuthenticated, HasTenant]

    def get(self, request):
        t_id = tenant_id(request)
        term = (request.query_params.get("term") or "").strip()
        if not term:
            return Response({"detail": "term is required"}, status=status.HTTP_400_BAD_REQUEST)

        bc = (
            Barcode.objects
            .filter(tenant_id=t_id, code=term)
            .select_related("product__category")
            .prefetch_related(Prefetch("product__barcodes", queryset=Barcode.objects.order_by("code")))
            .first()
        )
        if bc and bc.product:
            return Response(ProductReadSerializer(bc.product).data)

        p = (
            Product.objects
            .filter(tenant_id=t_id, sku__iexact=term)
            .select_related("category")
            .prefetch_related(_prefetch_barcodes_ordered())
            .first()
        )
        if p:
            return Response(ProductReadSerializer(p).data)

        return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)
    

class ProductImport(APIView):
    permission_classes = [IsAuthenticated, HasTenant]

    REQUIRED_HEADERS = {"name"}  # mínimo
    OPTIONAL_HEADERS = {"sku", "description", "unit", "price", "is_active", "category", "barcodes", "cost", "min_stock", "brand", "image_url"}
    MAX_ERRORS_RETURN = 50

    # Mapeo de headers en español / Fudo / otros POS → headers Pulstock
    HEADER_MAP = {
        "nombre": "name",
        "categoría": "category",
        "categoria": "category",
        "subcategoría": "category",      # fallback si no hay categoría
        "subcategoria": "category",
        "código": "sku",
        "codigo": "sku",
        "descripción": "description",
        "descripcion": "description",
        "precio": "price",
        "costo": "cost",
        "activo": "is_active",
        "stock": "min_stock",
        "proveedor": "brand",
        "marca": "brand",
    }

    @staticmethod
    def _strip_accents(s: str) -> str:
        """Remove accents/diacritics for header matching."""
        import unicodedata
        nfkd = unicodedata.normalize("NFKD", s)
        return "".join(c for c in nfkd if not unicodedata.combining(c))

    def _map_headers(self, fieldnames: list[str]) -> list[str]:
        """Traduce headers en español/Fudo a headers Pulstock."""
        mapped = []
        seen = set()
        for h in fieldnames:
            key = self._strip_accents((h or "").strip().lower())
            new_h = self.HEADER_MAP.get(key, h)
            # Evitar duplicados (ej: si ya hay "category" y viene "subcategoría")
            if new_h in seen:
                mapped.append(h)  # dejar original, se ignorará como unknown
            else:
                seen.add(new_h)
                mapped.append(new_h)
        return mapped

    def _excel_to_csv_text(self, raw_bytes: bytes, filename: str) -> str:
        """Convierte archivo Excel (.xls/.xlsx) a texto CSV en memoria."""
        import pandas as pd
        buf = io.BytesIO(raw_bytes)
        engine = "xlrd" if filename.endswith(".xls") else "openpyxl"
        df = pd.read_excel(buf, engine=engine, dtype=str)
        df = df.fillna("")
        out = io.StringIO()
        df.to_csv(out, index=False)
        return out.getvalue()

    def _is_excel(self, filename: str) -> bool:
        return filename.lower().endswith((".xls", ".xlsx"))

    def _sniff_dialect(self, text: str):
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ","
        return dialect

    def _normalize_bool(self, raw: str, default=True) -> bool:
        s = (raw or "").strip().lower()
        if s == "":
            return default
        truthy = {"1", "true", "t", "yes", "y", "si", "sí", "s", "on"}
        falsy = {"0", "false", "f", "no", "n", "off"}
        if s in truthy:
            return True
        if s in falsy:
            return False
        return default

    def _normalize_price(self, raw: str) -> Decimal:
        s = (raw or "").strip()
        if not s:
            return Decimal("0")

        s = s.replace("$", "").replace(" ", "")

        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")  # 1.234,56 -> 1234.56
            else:
                s = s.replace(",", "")  # 1,234.56 -> 1234.56
        elif "," in s and "." not in s:
            s = s.replace(",", ".")  # 1234,56 -> 1234.56

        try:
            return Decimal(s)
        except (InvalidOperation, ValueError):
            raise ValueError(f"Precio inválido: {raw}")

    def _clean_barcodes(self, raw: str):
        # acepta "code1,code2" o "code1|code2"
        raw = (raw or "").strip()
        if not raw:
            return []
        raw = raw.replace("|", ",")
        codes = [c.strip() for c in raw.split(",") if c.strip()]

        seen = set()
        clean = []
        for c in codes:
            if c in seen:
                continue
            seen.add(c)
            clean.append(c)

        if len(clean) > 30:
            raise ValueError("Too many barcodes (max 30).")
        return clean

    @transaction.atomic
    def post(self, request):
        t_id = getattr(request.user, "tenant_id", None)
        if not t_id:
            return Response({"detail": "User must have a tenant."}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "file is required (multipart/form-data key: file)"}, status=status.HTTP_400_BAD_REQUEST)

        MAX_CSV_SIZE = 5 * 1024 * 1024  # 5 MB
        if file.size > MAX_CSV_SIZE:
            return Response(
                {"detail": "Archivo demasiado grande. Máximo 5 MB."},
                status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            )

        raw = file.read()
        filename = getattr(file, "name", "") or ""

        if self._is_excel(filename):
            try:
                text = self._excel_to_csv_text(raw, filename)
            except Exception as e:
                return Response({"detail": f"Error al leer archivo Excel: {e}"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            try:
                text = raw.decode("utf-8-sig")  # soporta BOM
            except Exception:
                return Response({"detail": "Invalid encoding. Use UTF-8."}, status=status.HTTP_400_BAD_REQUEST)

        if not text.strip():
            return Response({"detail": "Empty file."}, status=status.HTTP_400_BAD_REQUEST)

        dialect = self._sniff_dialect(text)

        try:
            reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        except csv.Error as e:
            return Response({"detail": f"CSV parse error: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        if not reader.fieldnames:
            return Response({"detail": "CSV has no headers"}, status=status.HTTP_400_BAD_REQUEST)

        # Normaliza headers y aplica mapeo español → inglés
        fieldnames = [(h.strip() if isinstance(h, str) else h) for h in reader.fieldnames]
        fieldnames = self._map_headers(fieldnames)
        reader.fieldnames = fieldnames

        # Verificar límite de plan antes de procesar
        limit_check = _check_product_limit(t_id)
        if not limit_check["allowed"]:
            return Response(
                {
                    "detail": f"Has alcanzado el límite de {limit_check['limit']} productos de tu plan.",
                    "limit": limit_check["limit"],
                    "current": limit_check["current"],
                },
                status=status.HTTP_402_PAYMENT_REQUIRED,
            )

        headers = set([h for h in fieldnames if h])
        missing = self.REQUIRED_HEADERS - headers
        if missing:
            return Response(
                {"detail": f"Missing required headers: {sorted(list(missing))}", "headers": fieldnames},
                status=status.HTTP_400_BAD_REQUEST,
            )

        unknown = headers - (self.REQUIRED_HEADERS | self.OPTIONAL_HEADERS)

        created = 0
        updated = 0
        skipped = 0
        errors = []

        # cache categorías por nombre (lower)
        cat_cache = {c.name.strip().lower(): c for c in Category.objects.filter(tenant_id=t_id)}

        expected_keys = set(fieldnames)

        MAX_ROWS = 5000
        for i, row in enumerate(reader, start=2):
            if i - 1 > MAX_ROWS:
                errors.append({"line": i, "error": f"Máximo {MAX_ROWS} filas. Filas restantes ignoradas."})
                break
            sid = transaction.savepoint()
            try:
                # Fila malformada: columnas sobrantes
                row_keys = set(row.keys())
                if None in row_keys:
                    errors.append({
                        "line": i,
                        "error": "Fila mal formada: columnas sobrantes (probable comilla mal cerrada o delimiter incorrecto).",
                        "extra": row.get(None),
                        "row": {k: v for k, v in row.items() if k is not None},
                    })
                    skipped += 1
                    continue

                # Columnas desalineadas
                if row_keys != expected_keys:
                    errors.append({
                        "line": i,
                        "error": "Fila mal formada: columnas desalineadas (no coincide con header).",
                        "row_keys": sorted(list(row_keys)),
                        "expected_keys": sorted(list(expected_keys)),
                        "row": row,
                    })
                    skipped += 1
                    continue

                # Parse
                name = (row.get("name") or "").strip()
                if not name:
                    skipped += 1
                    continue

                sku = (row.get("sku") or "").strip()
                desc = (row.get("description") or "").strip()
                unit = ((row.get("unit") or "").strip() or "UN").upper()

                price_raw = (row.get("price") or "").strip()
                is_active_raw = (row.get("is_active") or "").strip()
                cat_name = (row.get("category") or "").strip()
                barcodes_raw = (row.get("barcodes") or "").strip()

                # new fields
                cost_raw = (row.get("cost") or "").strip()
                min_stock_raw = (row.get("min_stock") or "").strip()
                brand = (row.get("brand") or "").strip()
                image_url = (row.get("image_url") or "").strip()

                price = self._normalize_price(price_raw)
                is_active = self._normalize_bool(is_active_raw, default=True)
                cost = self._normalize_price(cost_raw) if cost_raw else Decimal("0")
                min_stock = self._normalize_price(min_stock_raw) if min_stock_raw else Decimal("0")

                # category (create si no existe)
                category = None
                if cat_name:
                    key = cat_name.strip().lower()
                    category = cat_cache.get(key)
                    if not category:
                        category = Category.objects.create(tenant_id=t_id, name=cat_name.strip(), code="")
                        cat_cache[key] = category

                # upsert
                product = None
                if sku:
                    product = Product.objects.filter(tenant_id=t_id, sku__iexact=sku).first()
                if not product:
                    product = Product.objects.filter(tenant_id=t_id, name__iexact=name).first()

                # Auto-resolve unit_obj FK from unit string
                unit_obj = Unit.objects.filter(
                    tenant_id=t_id, code=unit, is_active=True
                ).first()

                if product:
                    product.name = name
                    product.sku = sku
                    product.description = desc
                    product.unit = unit
                    product.price = price
                    product.is_active = is_active
                    product.category = category
                    if unit_obj:
                        product.unit_obj = unit_obj
                    if cost_raw:
                        product.cost = cost
                    if min_stock_raw:
                        product.min_stock = min_stock
                    if brand:
                        product.brand = brand
                    if image_url:
                        product.image_url = image_url
                    product.save()
                    updated += 1
                else:
                    product = Product.objects.create(
                        tenant_id=t_id,
                        name=name,
                        sku=sku,
                        description=desc,
                        unit=unit,
                        unit_obj=unit_obj,
                        price=price,
                        is_active=is_active,
                        category=category,
                        cost=cost,
                        min_stock=min_stock,
                        brand=brand,
                        image_url=image_url,
                    )
                    created += 1

                # barcodes (reemplazo total si la columna viene)
                if "barcodes" in expected_keys:
                    clean = self._clean_barcodes(barcodes_raw)

                    if clean:
                        collisions = (
                            Barcode.objects
                            .filter(tenant_id=t_id, code__in=clean)
                            .exclude(product=product)
                            .values_list("code", flat=True)
                            .distinct()
                        )
                        collisions = list(collisions)
                        if collisions:
                            raise ValueError(f"Barcode(s) ya usados por otro producto: {', '.join(collisions[:10])}")

                    Barcode.objects.filter(tenant_id=t_id, product=product).delete()

                    if clean:
                        Barcode.objects.bulk_create([
                            Barcode(tenant_id=t_id, product=product, code=code)
                            for code in clean
                        ])

                transaction.savepoint_commit(sid)
            except Exception as e:
                transaction.savepoint_rollback(sid)
                errors.append({"line": i, "error": str(e), "row": row})
                skipped += 1
                continue

        return Response(
            {
                "delimiter_detected": getattr(dialect, "delimiter", ","),
                "headers": fieldnames,
                "unknown_headers": sorted(list(unknown)) if unknown else [],
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors_count": len(errors),
                "errors": errors[: self.MAX_ERRORS_RETURN],
            },
            status=status.HTTP_200_OK,
        )
    
class CategoryTree(generics.ListAPIView):
    """
    GET /api/catalog/categories/tree/
    Devuelve árbol completo para selectores de UI.
    """
    permission_classes = [IsAuthenticated, HasTenant]
    serializer_class = CategorySerializer

    def get_queryset(self):
        return (
            Category.objects
            .filter(tenant_id=tenant_id(self.request), parent__isnull=True)
            .prefetch_related("children")
            .order_by("name")
        )


# ══════════════════════════════════════════════════════════════════════════════
# SAMPLE EXCEL DOWNLOAD
# ══════════════════════════════════════════════════════════════════════════════

class ProductImportSample(APIView):
    """GET /catalog/products/import-sample/ → descargar Excel de ejemplo."""
    permission_classes = [IsAuthenticated, HasTenant]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse

        wb = Workbook()

        # ── Hoja 1: Ejemplo de productos ──
        ws = wb.active
        ws.title = "Productos"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        headers = ["name", "sku", "description", "unit", "price", "cost",
                    "is_active", "category", "barcodes", "min_stock", "brand"]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        examples = [
            ["Arroz Grano Largo 1kg", "ARR-001", "Arroz blanco premium", "KG",
             2500, 1500, "si", "Abarrotes", "7801234567890", 5, "Marca A"],
            ["Aceite Oliva 500ml", "ACE-002", "Aceite extra virgen", "UN",
             8500, 6000, "si", "Condimentos", "7801234567891|7801234567892", 2, "Marca B"],
            ["Café Molido 250g", "CAF-003", "", "UN",
             5500, 4000, "si", "Bebidas", "7801234567893", 10, ""],
            ["Servilletas x100", "SER-004", "Pack 100 unidades", "PAQ",
             1200, 700, "si", "Limpieza", "", 20, ""],
            ["Harina 1kg", "HAR-005", "Harina sin preparar", "KG",
             1800, 1100, "si", "Abarrotes", "7801234567894", 8, "Marca C"],
        ]

        for row_idx, row_data in enumerate(examples, 2):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

        # Ajustar anchos
        widths = [25, 12, 25, 6, 10, 10, 10, 15, 30, 10, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        # ── Hoja 2: Instrucciones ──
        ws2 = wb.create_sheet("Instrucciones")
        ws2.sheet_properties.tabColor = "10B981"

        instructions = [
            ["Campo", "Requerido", "Descripción", "Ejemplo"],
            ["name", "Sí", "Nombre del producto", "Arroz Grano Largo 1kg"],
            ["sku", "No", "Código único del producto. Si existe, se actualiza en vez de crear", "ARR-001"],
            ["description", "No", "Descripción del producto", "Arroz blanco premium"],
            ["unit", "No", "Unidad de medida: UN, KG, GR, LT, ML, MT, CM, CAJA, PAQ, DOC (defecto: UN)", "KG"],
            ["price", "No", "Precio de venta (sin símbolo $). Usa punto como decimal", "2500"],
            ["cost", "No", "Costo del producto (sin símbolo $)", "1500"],
            ["is_active", "No", "Activo: si/no, 1/0, true/false (defecto: si)", "si"],
            ["category", "No", "Nombre de la categoría. Si no existe, se crea automáticamente", "Abarrotes"],
            ["barcodes", "No", "Código(s) de barra. Separar múltiples con | (pipe)", "7801234567890|7801234567891"],
            ["min_stock", "No", "Stock mínimo para alertas", "5"],
            ["brand", "No", "Marca del producto", "Marca A"],
        ]

        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

        ws2.column_dimensions["A"].width = 15
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 60
        ws2.column_dimensions["D"].width = 30

        # Notas
        notes_start = len(instructions) + 2
        notes = [
            "NOTAS IMPORTANTES:",
            "• Guardar como CSV UTF-8 antes de importar (Archivo → Guardar como → CSV UTF-8)",
            "• Si el SKU ya existe, el producto se ACTUALIZA (no se duplica)",
            "• Si el nombre ya existe (sin SKU), también se actualiza",
            "• Las categorías se crean automáticamente si no existen",
            "• Los códigos de barra deben ser únicos en todo el sistema",
            "• Máximo 30 códigos de barra por producto (separar con |)",
            "• Los precios y costos usan punto como decimal (ej: 2500.50)",
        ]
        for i, note in enumerate(notes):
            cell = ws2.cell(row=notes_start + i, column=1, value=note)
            if i == 0:
                cell.font = Font(bold=True, size=12, color="DC2626")
            else:
                cell.font = Font(size=10, color="6B7280")
            ws2.merge_cells(start_row=notes_start + i, start_column=1,
                            end_row=notes_start + i, end_column=4)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="ejemplo_productos.xlsx"'
        wb.save(response)
        return response


# ══════════════════════════════════════════════════════════════════════════════
# RECIPE BULK IMPORT
# ══════════════════════════════════════════════════════════════════════════════

class RecipeImport(APIView):
    """
    POST /catalog/recipes/import-csv/ → carga masiva de recetas desde CSV.

    Formato CSV:
        product_sku, product_name, ingredient_sku, ingredient_name, qty

    Lógica:
        - Identifica producto padre por SKU o nombre
        - Identifica ingrediente por SKU o nombre
        - Agrupa líneas por producto padre → crea/reemplaza receta
    """
    permission_classes = [IsAuthenticated, HasTenant]

    def post(self, request):
        t_id = tenant_id(request)
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "file is required"}, status=400)

        if file.size > 5 * 1024 * 1024:
            return Response(
                {"detail": "Archivo demasiado grande (máximo 5 MB)"},
                status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            )

        try:
            raw = file.read().decode("utf-8-sig")
        except UnicodeDecodeError:
            return Response({"detail": "Encoding inválido. Usa UTF-8."}, status=400)

        if not raw.strip():
            return Response({"detail": "Archivo vacío."}, status=400)

        # Detect delimiter
        try:
            dialect = csv.Sniffer().sniff(raw[:2048], delimiters=",;")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","

        reader = csv.DictReader(io.StringIO(raw), delimiter=delimiter)
        if not reader.fieldnames:
            return Response({"detail": "CSV sin encabezados."}, status=400)

        # Normalize headers
        clean_headers = {h.strip().lower().replace(" ", "_"): h for h in reader.fieldnames}

        # Validate required headers
        has_product = "product_sku" in clean_headers or "product_name" in clean_headers
        has_ingredient = "ingredient_sku" in clean_headers or "ingredient_name" in clean_headers
        has_qty = "qty" in clean_headers

        if not has_product or not has_ingredient or not has_qty:
            return Response({
                "detail": "Faltan columnas requeridas. Necesitas: "
                          "(product_sku o product_name), "
                          "(ingredient_sku o ingredient_name), qty"
            }, status=400)

        # Cache products by SKU and name
        products_by_sku = {}
        products_by_name = {}
        for p in Product.objects.filter(tenant_id=t_id, is_active=True):
            if p.sku:
                products_by_sku[p.sku.strip().upper()] = p
            products_by_name[p.name.strip().lower()] = p

        # Parse rows → group by parent product
        recipes_data = {}  # product_id -> [(ingredient, qty)]
        errors = []

        for line_num, row in enumerate(reader, start=2):
            # Normalize keys
            norm = {}
            for k, v in row.items():
                nk = k.strip().lower().replace(" ", "_") if k else ""
                norm[nk] = (v or "").strip()

            # Find parent product
            parent = None
            p_sku = norm.get("product_sku", "").strip()
            p_name = norm.get("product_name", "").strip()

            if p_sku:
                parent = products_by_sku.get(p_sku.upper())
            if not parent and p_name:
                parent = products_by_name.get(p_name.lower())

            if not parent:
                errors.append({
                    "line": line_num,
                    "error": f"Producto padre no encontrado: SKU='{p_sku}' nombre='{p_name}'",
                })
                continue

            # Find ingredient
            ingredient = None
            i_sku = norm.get("ingredient_sku", "").strip()
            i_name = norm.get("ingredient_name", "").strip()

            if i_sku:
                ingredient = products_by_sku.get(i_sku.upper())
            if not ingredient and i_name:
                ingredient = products_by_name.get(i_name.lower())

            if not ingredient:
                errors.append({
                    "line": line_num,
                    "error": f"Ingrediente no encontrado: SKU='{i_sku}' nombre='{i_name}'",
                })
                continue

            if ingredient.id == parent.id:
                errors.append({
                    "line": line_num,
                    "error": f"Un producto no puede ser ingrediente de sí mismo: '{parent.name}'",
                })
                continue

            # Parse qty
            qty_raw = norm.get("qty", "").replace(",", ".")
            try:
                qty = Decimal(qty_raw)
                if qty <= 0:
                    raise ValueError()
            except (InvalidOperation, ValueError):
                errors.append({
                    "line": line_num,
                    "error": f"Cantidad inválida: '{norm.get('qty', '')}'",
                })
                continue

            recipes_data.setdefault(parent.id, {"product": parent, "lines": [], "_seen_ings": set()})
            if ingredient.id in recipes_data[parent.id]["_seen_ings"]:
                errors.append({
                    "line": line_num,
                    "error": f"Ingrediente duplicado para '{parent.name}': '{ingredient.name}' — se usa la primera aparición.",
                })
                continue
            recipes_data[parent.id]["_seen_ings"].add(ingredient.id)
            recipes_data[parent.id]["lines"].append({
                "ingredient": ingredient,
                "qty": qty,
            })

        # Save recipes
        created = 0
        updated = 0

        with transaction.atomic():
            for pid, rdata in recipes_data.items():
                product = rdata["product"]
                lines = rdata["lines"]

                existing = Recipe.objects.filter(
                    tenant_id=t_id, product=product
                ).first()

                recipe, _ = Recipe.objects.update_or_create(
                    tenant_id=t_id, product=product,
                    defaults={"is_active": True},
                )
                RecipeLine.objects.filter(recipe=recipe).delete()
                RecipeLine.objects.bulk_create([
                    RecipeLine(
                        tenant_id=t_id,
                        recipe=recipe,
                        ingredient_id=line["ingredient"].id,
                        qty=line["qty"],
                    )
                    for line in lines
                ])

                if existing:
                    updated += 1
                else:
                    created += 1

        return Response({
            "created": created,
            "updated": updated,
            "errors_count": len(errors),
            "errors": errors[:50],
        })


class RecipeImportSample(APIView):
    """GET /catalog/recipes/import-sample/ → descargar Excel de ejemplo para recetas."""
    permission_classes = [IsAuthenticated, HasTenant]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = "Recetas"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        headers = ["product_sku", "product_name", "ingredient_sku", "ingredient_name", "qty"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        examples = [
            ["PIZZA-001", "Pizza Margarita", "MASA-001", "Masa para pizza", "1"],
            ["PIZZA-001", "Pizza Margarita", "SAL-001", "Salsa de tomate", "0.15"],
            ["PIZZA-001", "Pizza Margarita", "QUES-001", "Queso mozzarella", "0.2"],
            ["PIZZA-001", "Pizza Margarita", "ALB-001", "Albahaca fresca", "0.01"],
            ["SAND-001", "Sándwich Completo", "PAN-001", "Pan de molde", "2"],
            ["SAND-001", "Sándwich Completo", "JAM-001", "Jamón laminado", "0.08"],
            ["SAND-001", "Sándwich Completo", "QUES-001", "Queso mozzarella", "0.06"],
            ["SAND-001", "Sándwich Completo", "TOMATE", "Tomate", "0.1"],
            ["CAFE-LAT", "Café Latte", "CAFE-G", "Café en grano", "0.02"],
            ["CAFE-LAT", "Café Latte", "LECHE", "Leche entera", "0.25"],
        ]

        for row_idx, row_data in enumerate(examples, 2):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

        widths = [15, 22, 15, 22, 8]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        # Instrucciones
        ws2 = wb.create_sheet("Instrucciones")
        ws2.sheet_properties.tabColor = "7C3AED"

        instructions = [
            ["Campo", "Requerido", "Descripción", "Ejemplo"],
            ["product_sku", "Uno de los dos", "SKU del producto padre (el que se vende)", "PIZZA-001"],
            ["product_name", "Uno de los dos", "Nombre del producto padre (si no hay SKU)", "Pizza Margarita"],
            ["ingredient_sku", "Uno de los dos", "SKU del ingrediente a descontar", "MASA-001"],
            ["ingredient_name", "Uno de los dos", "Nombre del ingrediente (si no hay SKU)", "Masa para pizza"],
            ["qty", "Sí", "Cantidad del ingrediente por 1 unidad vendida del producto padre", "0.15"],
        ]

        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 15
        ws2.column_dimensions["C"].width = 55
        ws2.column_dimensions["D"].width = 20

        notes_start = len(instructions) + 2
        notes = [
            "NOTAS IMPORTANTES:",
            "• Guardar como CSV UTF-8 antes de importar",
            "• Cada fila es UNA línea de receta (producto → ingrediente → cantidad)",
            "• Un producto con 3 ingredientes = 3 filas con el mismo product_sku",
            "• Si el producto ya tiene receta, se REEMPLAZA completamente",
            "• Tanto el producto padre como el ingrediente deben existir previamente",
            "• La cantidad (qty) es por 1 unidad vendida del producto padre",
            "• Ejemplo: si 1 pizza usa 200g de queso, qty = 0.2 (en KG)",
            "• Puedes usar SKU o nombre para identificar productos, pero SKU es más confiable",
        ]
        for i, note in enumerate(notes):
            cell = ws2.cell(row=notes_start + i, column=1, value=note)
            if i == 0:
                cell.font = Font(bold=True, size=12, color="DC2626")
            else:
                cell.font = Font(size=10, color="6B7280")
            ws2.merge_cells(start_row=notes_start + i, start_column=1,
                            end_row=notes_start + i, end_column=4)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="ejemplo_recetas.xlsx"'
        wb.save(response)
        return response


class ProductRecipeView(APIView):
    """
    GET  /catalog/products/{pk}/recipe/  → obtener receta (404 si no tiene)
    POST /catalog/products/{pk}/recipe/  → crear o reemplazar receta completa
    DELETE /catalog/products/{pk}/recipe/ → eliminar receta
    """
    permission_classes = [IsAuthenticated, HasTenant]

    def _get_product(self, pk, t_id):
        try:
            return Product.objects.get(pk=pk, tenant_id=t_id)
        except Product.DoesNotExist:
            return None

    def get(self, request, pk):
        t_id = tenant_id(request)
        product = self._get_product(pk, t_id)
        if not product:
            return Response({"detail": "Producto no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        try:
            recipe = Recipe.objects.prefetch_related("lines__ingredient__unit_obj", "lines__unit").get(
                product=product, tenant_id=t_id
            )
        except Recipe.DoesNotExist:
            return Response({"detail": "Este producto no tiene receta."}, status=status.HTTP_404_NOT_FOUND)
        return Response(RecipeReadSerializer(recipe).data)

    @transaction.atomic
    def post(self, request, pk):
        t_id = tenant_id(request)
        product = self._get_product(pk, t_id)
        if not product:
            return Response({"detail": "Producto no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        ser = RecipeWriteSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        # Validar ingredientes pertenecen al tenant y son productos activos
        line_data = data["lines"]
        ingredient_ids = [l["ingredient_id"] for l in line_data]
        ingredients = {
            p.id: p for p in
            Product.objects.filter(tenant_id=t_id, id__in=ingredient_ids, is_active=True)
        }
        missing = set(ingredient_ids) - set(ingredients.keys())
        if missing:
            return Response(
                {"detail": f"Ingredientes no válidos o inactivos: {sorted(missing)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        # Ningún ingrediente puede ser el mismo producto
        if product.id in ingredients:
            return Response(
                {"detail": "Un producto no puede ser ingrediente de su propia receta."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Detección de ciclos transitivos (A→B→C→A)
        # BFS: desde cada ingrediente, seguir sus recetas buscando si alguno
        # eventualmente usa product.id como ingrediente
        def _has_cycle(target_id):
            queue = list(ingredient_ids)
            seen = set(queue)
            while queue:
                # Para los nodos en la cola, buscar qué ingredientes usan en SUS recetas
                batch = RecipeLine.objects.filter(
                    tenant_id=t_id, recipe__product_id__in=queue,
                ).values_list("ingredient_id", flat=True)
                queue = []
                for child_id in batch:
                    if child_id == target_id:
                        return True
                    if child_id not in seen:
                        seen.add(child_id)
                        queue.append(child_id)
            return False

        if _has_cycle(product.id):
            return Response(
                {"detail": "Dependencia circular detectada: un ingrediente ya usa este producto en su receta."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validar unidades de líneas (si se especificaron)
        unit_ids_in_lines = [l["unit_id"] for l in line_data if l.get("unit_id")]
        units_by_id = {}
        if unit_ids_in_lines:
            units_by_id = {
                u.id: u for u in
                Unit.objects.filter(tenant_id=t_id, id__in=unit_ids_in_lines, is_active=True)
            }
            missing_units = set(unit_ids_in_lines) - set(units_by_id.keys())
            if missing_units:
                return Response(
                    {"detail": f"Unidades no válidas: {sorted(missing_units)}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            # Pre-fetch ingredient units to avoid N+1
            ing_unit_ids = {ing.unit_obj_id for ing in ingredients.values() if ing.unit_obj_id}
            ing_units_by_id = {u.id: u for u in Unit.objects.filter(id__in=ing_unit_ids)} if ing_unit_ids else {}

            # Validar familia compatible con el ingrediente
            for l in line_data:
                uid = l.get("unit_id")
                if not uid:
                    continue
                line_unit = units_by_id[uid]
                ing = ingredients[l["ingredient_id"]]
                if ing.unit_obj_id:
                    ing_unit = ing_units_by_id.get(ing.unit_obj_id)
                    if ing_unit and line_unit.family != ing_unit.family:
                        return Response(
                            {"detail": f"Unidad {line_unit.code} ({line_unit.get_family_display()}) "
                                       f"no es compatible con {ing.name} ({ing_unit.code})."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

        # Crear o reemplazar receta
        recipe, _ = Recipe.objects.update_or_create(
            tenant_id=t_id, product=product,
            defaults={"is_active": data["is_active"], "notes": data.get("notes", "")},
        )
        RecipeLine.objects.filter(recipe=recipe).delete()
        RecipeLine.objects.bulk_create([
            RecipeLine(
                tenant_id=t_id,
                recipe=recipe,
                ingredient_id=l["ingredient_id"],
                qty=l["qty"],
                unit_id=l.get("unit_id"),
            )
            for l in line_data
        ])

        recipe.refresh_from_db()
        recipe_out = Recipe.objects.prefetch_related("lines__ingredient", "lines__unit").get(pk=recipe.pk)
        return Response(RecipeReadSerializer(recipe_out).data, status=status.HTTP_200_OK)

    @transaction.atomic
    def delete(self, request, pk):
        t_id = tenant_id(request)
        product = self._get_product(pk, t_id)
        if not product:
            return Response({"detail": "Producto no encontrado."}, status=status.HTTP_404_NOT_FOUND)
        deleted, _ = Recipe.objects.filter(product=product, tenant_id=t_id).delete()
        if not deleted:
            return Response({"detail": "Este producto no tiene receta."}, status=status.HTTP_404_NOT_FOUND)
        return Response(status=status.HTTP_204_NO_CONTENT)


class UnitList(generics.ListAPIView):
    """GET /catalog/units/ → lista de unidades del tenant."""
    permission_classes = [IsAuthenticated, HasTenant]
    serializer_class = UnitSerializer
    pagination_class = None

    def get_queryset(self):
        return Unit.objects.filter(
            tenant_id=tenant_id(self.request), is_active=True,
        ).order_by("family", "code")


# ══════════════════════════════════════════════════════════════════════════════
# PRICE LIST — Vista y actualización masiva de precios
# ══════════════════════════════════════════════════════════════════════════════

class PriceListView(APIView):
    """GET /catalog/products/prices/ — Lista de productos con info de precios."""
    permission_classes = [IsAuthenticated, HasTenant]

    PAGE_SIZE = 50

    def get(self, request):
        qs = Product.objects.filter(
            tenant_id=tenant_id(request), is_active=True,
        ).select_related("category").order_by("name")

        q = request.query_params.get("q", "").strip()
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(sku__icontains=q))

        cat_id = request.query_params.get("category_id")
        if cat_id:
            from catalog.models import Category
            if not Category.objects.filter(id=cat_id, tenant_id=tenant_id(request)).exists():
                return Response({"detail": "Categoría no encontrada."}, status=404)
            qs = qs.filter(category_id=cat_id)

        # Paginación real
        total = qs.count()
        try:
            page = max(1, int(request.query_params.get("page", 1)))
        except (ValueError, TypeError):
            page = 1
        start = (page - 1) * self.PAGE_SIZE
        page_qs = qs[start:start + self.PAGE_SIZE]

        data = []
        for p in page_qs:
            cost = p.cost or Decimal("0")
            margin = ((p.price - cost) / p.price * 100) if p.price and p.price > 0 else Decimal("0")
            data.append({
                "id": p.id,
                "sku": p.sku,
                "name": p.name,
                "category_name": p.category.name if p.category else None,
                "category_id": p.category_id,
                "cost": str(cost),
                "price": str(p.price),
                "margin_pct": str(margin.quantize(Decimal("0.1"))),
            })
        return Response({
            "results": data,
            "total": total,
            "page": page,
            "page_size": self.PAGE_SIZE,
        })


class PriceBulkUpdateView(APIView):
    """POST /catalog/products/prices/bulk/ — Actualización masiva de precios."""
    permission_classes = [IsAuthenticated, HasTenant, IsManager]

    def post(self, request):
        tid = tenant_id(request)
        updates = request.data.get("updates")
        rule = request.data.get("rule")

        if updates:
            # Modo individual: [{ product_id, price }]
            count = 0
            for item in updates:
                try:
                    pid = item["product_id"]
                    new_price = Decimal(str(item["price"]))
                    if new_price < 0:
                        continue
                    Product.objects.filter(id=pid, tenant_id=tid).update(price=new_price)
                    count += 1
                except (KeyError, InvalidOperation):
                    continue
            return Response({"updated": count})

        elif rule:
            # Modo masivo: { type, value, direction } + filter
            try:
                rtype = rule["type"]  # "pct" o "amt"
                value = Decimal(str(rule["value"]))
                direction = rule["direction"]  # "increase" o "decrease"
            except (KeyError, InvalidOperation):
                return Response({"detail": "Regla inválida."}, status=400)

            if value <= 0:
                return Response({"detail": "El valor debe ser mayor a 0."}, status=400)

            qs = Product.objects.filter(tenant_id=tid, is_active=True)
            filt = request.data.get("filter", {})
            if filt.get("category_id"):
                from catalog.models import Category
                if not Category.objects.filter(id=filt["category_id"], tenant_id=tid).exists():
                    return Response({"detail": "Categoría no encontrada."}, status=404)
                qs = qs.filter(category_id=filt["category_id"])
            if filt.get("product_ids"):
                # Validate all product_ids belong to this tenant
                valid_count = Product.objects.filter(
                    id__in=filt["product_ids"], tenant_id=tid
                ).count()
                if valid_count != len(filt["product_ids"]):
                    return Response({"detail": "Algunos productos no pertenecen a tu negocio."}, status=400)
                qs = qs.filter(id__in=filt["product_ids"])

            if rtype == "pct":
                if direction == "decrease" and value >= 100:
                    return Response({"detail": "No se puede reducir 100% o más."}, status=400)
                factor = Decimal("1") + (value / Decimal("100") if direction == "increase" else -value / Decimal("100"))
                count = qs.update(price=Greatest(F("price") * factor, Decimal("0")))
            else:  # amt
                if direction == "increase":
                    count = qs.update(price=F("price") + value)
                else:
                    # Evitar precios negativos: solo actualizar donde precio >= valor
                    count = qs.filter(price__gte=value).update(price=F("price") - value)

            return Response({"updated": count})

        return Response({"detail": "Envía 'updates' o 'rule'."}, status=400)