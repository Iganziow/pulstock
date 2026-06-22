"""
Tests para los fixes del módulo de reportes (jun 2026).

Cubren:
- #38 audit-trail: ref_type NULL ya no crashea (TypeError en sorted() →
  "Dato con formato incorrecto"). Este es el bug que reportó el usuario.
- #40 sales-summary: la serie diaria respeta el filtro de categoría.
- #42 validación de rango de fechas invertido → 400 claro (no vacío silencioso).
- #43 inventory-health: excluye productos inactivos.
- #44 toma física: requiere rol inventario/manager (cajero → 403).
- #39 export: respeta el filtro de categoría (no descarga toda la tienda).
"""
import pytest
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from rest_framework.test import APIClient

from core.models import User
from catalog.models import Category, Product
from inventory.models import StockItem, StockMove
from sales.services import create_sale

_sub = pytest.mark.usefixtures("forecast_subscription")


def _stock(tenant, warehouse, product, on_hand, avg_cost="100"):
    return StockItem.objects.create(
        tenant=tenant, warehouse=warehouse, product=product,
        on_hand=Decimal(str(on_hand)), avg_cost=Decimal(str(avg_cost)),
        stock_value=Decimal(str(on_hand)) * Decimal(str(avg_cost)),
    )


def _sell(tenant, store, warehouse, owner, lines):
    """lines: [(product, qty, unit_price), ...]"""
    total = sum(Decimal(str(q)) * Decimal(str(pr)) for _, q, pr in lines)
    return create_sale(
        user=owner, tenant_id=tenant.id, store_id=store.id, warehouse_id=warehouse.id,
        lines_in=[{"product_id": p.id, "qty": str(q), "unit_price": str(pr)} for p, q, pr in lines],
        payments_in=[{"method": "cash", "amount": str(total)}],
    )


# ── #38 audit-trail con ref_type NULL ────────────────────────────────────────
@_sub
@pytest.mark.django_db
class TestAuditTrailNullRefType:
    URL = "/api/reports/audit-trail/"

    def test_null_ref_type_does_not_crash(self, api_client, tenant, warehouse, product):
        # Movimiento con ref_type NULL (caso real Marbrava: ajustes / data vieja).
        StockMove.objects.create(
            tenant=tenant, warehouse=warehouse, product=product,
            move_type="OUT", ref_type=None, qty=Decimal("1"),
            cost_snapshot=Decimal("0"), value_delta=Decimal("0"),
        )
        # ...y uno con ref_type string, para forzar la comparación None < str
        # que rompía sorted() en get_audit_trail.
        StockMove.objects.create(
            tenant=tenant, warehouse=warehouse, product=product,
            move_type="IN", ref_type="RECEIVE", qty=Decimal("5"),
            cost_snapshot=Decimal("0"), value_delta=Decimal("0"),
        )
        resp = api_client.get(self.URL)
        # Antes: 400 "Dato con formato incorrecto" (TypeError en sorted()).
        assert resp.status_code == 200, resp.content[:300]
        data = resp.json()
        assert "RECEIVE" in data["ref_types"]
        assert None not in data["ref_types"]
        assert data["meta"]["total"] >= 2


# ── #42 rango de fechas invertido ────────────────────────────────────────────
@_sub
@pytest.mark.django_db
def test_inverted_date_range_returns_400(api_client):
    resp = api_client.get(
        "/api/reports/sales-summary/",
        {"date_from": "2026-06-22", "date_to": "2026-06-15"},
    )
    assert resp.status_code == 400, resp.content[:300]


# ── #40 sales-summary: diario respeta categoría ──────────────────────────────
@pytest.mark.django_db
def test_sales_summary_daily_respects_category(tenant, store, warehouse, owner):
    from reports.services import get_sales_summary
    cat_a = Category.objects.create(tenant=tenant, name="Cafés")
    cat_b = Category.objects.create(tenant=tenant, name="Postres")
    p_a = Product.objects.create(tenant=tenant, category=cat_a, name="Latte", price=Decimal("3000"), is_active=True)
    p_b = Product.objects.create(tenant=tenant, category=cat_b, name="Torta", price=Decimal("5000"), is_active=True)
    _stock(tenant, warehouse, p_a, 100)
    _stock(tenant, warehouse, p_b, 100)
    _sell(tenant, store, warehouse, owner, [(p_a, 2, 3000), (p_b, 1, 5000)])

    data = get_sales_summary(tenant.id, store.id, category_id=cat_a.id)
    kpi_rev = Decimal(data["kpis"]["total_revenue"])
    daily_rev = sum(Decimal(d["revenue"]) for d in data["daily"])
    # Solo Latte (2×3000 = 6000), NO la torta.
    assert kpi_rev == Decimal("6000"), data["kpis"]
    assert daily_rev == kpi_rev, (
        f"daily {daily_rev} != kpi {kpi_rev}: la serie diaria ignoraba la categoría")


# ── #43 inventory-health excluye inactivos ───────────────────────────────────
@_sub
@pytest.mark.django_db
def test_inventory_health_excludes_inactive(api_client, tenant, warehouse):
    Product.objects.create(tenant=tenant, name="Activo IH", price=Decimal("100"), is_active=True)
    inactivo = Product.objects.create(tenant=tenant, name="Inactivo IH", price=Decimal("100"), is_active=False)
    _stock(tenant, warehouse, Product.objects.get(name="Activo IH"), 5)
    _stock(tenant, warehouse, inactivo, 5)

    resp = api_client.get("/api/reports/inventory-health/")
    assert resp.status_code == 200, resp.content[:300]
    data = resp.json()
    # Solo el producto activo entra al universo del reporte; el inactivo se excluye.
    assert data["summary"]["total_products"] == 1, data["summary"]
    # Y el inactivo no aparece en ninguna sección de riesgo.
    secs = ("zero_stock", "dead_stock", "below_minimum", "overstock", "discrepancies")
    names = [
        (r.get("product_name") or r.get("name"))
        for sec in secs for r in data.get(sec, [])
    ]
    assert "Inactivo IH" not in names, f"un producto INACTIVO no debe aparecer: {names}"


# ── #44 toma física requiere rol inventario/manager ──────────────────────────
@pytest.mark.django_db
def test_toma_fisica_forbidden_for_cashier(tenant, store, warehouse):
    cashier = User.objects.create(
        username="cash_tf", tenant=tenant, active_store=store, role=User.Role.CASHIER,
    )
    cashier.set_password("x")
    cashier.save()
    c = APIClient()
    c.force_authenticate(user=cashier)
    resp = c.get("/api/reports/inventory-count-sheet/")
    assert resp.status_code == 403, f"un cajero NO debería ver la toma física: {resp.status_code}"


# ── #39 export respeta filtro de categoría ───────────────────────────────────
@_sub
@pytest.mark.django_db
def test_sales_export_respects_category(api_client, tenant, store, warehouse, owner):
    cat_a = Category.objects.create(tenant=tenant, name="ExpA")
    cat_b = Category.objects.create(tenant=tenant, name="ExpB")
    p_a = Product.objects.create(tenant=tenant, category=cat_a, name="ProdExpA", price=Decimal("3000"), is_active=True)
    p_b = Product.objects.create(tenant=tenant, category=cat_b, name="ProdExpB", price=Decimal("5000"), is_active=True)
    _stock(tenant, warehouse, p_a, 100)
    _stock(tenant, warehouse, p_b, 100)
    _sell(tenant, store, warehouse, owner, [(p_a, 2, 3000), (p_b, 1, 5000)])

    r = api_client.get("/api/reports/sales-summary/export/", {"category_id": cat_a.id})
    assert r.status_code == 200, r.content[:200]
    ws = load_workbook(BytesIO(r.content)).active
    rows = list(ws.iter_rows(values_only=True))[1:]  # sin header
    ingresos = sum(float(rr[2] or 0) for rr in rows)  # col "Ingresos"
    # Solo CatA (6000), NO el total de la tienda (11000).
    assert ingresos == 6000, f"el export no respetó el filtro de categoría: ingresos={ingresos}"
