"""
Fix auditoría (09/06/26): los exports de reportes leían claves equivocadas
del dict que devuelven los servicios → salían VACÍOS o en $0:
  - Ventas Excel/PDF: leía total/profit; el servicio da revenue/cost.
  - Stock valorizado: leía warehouses/items; el servicio da una lista plana "results".
  - ABC PDF + email semanal: leía "items"; el servicio da "results".
  - Mermas: leía date/cost; details da created_at/value_lost (+ ahora sku).
  - Audit trail export: faltaba filtro warehouse__store_id → fuga cross-store.

Tests de integración: con datos sembrados, el export trae filas con valores
reales (no 0 / no vacío) y el audit trail respeta el scope de tienda.
"""
import pytest
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from core.models import Warehouse
from stores.models import Store
from catalog.models import Product
from inventory.models import StockItem, StockMove
from sales.services import create_sale


def _xlsx_rows(response):
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


@pytest.fixture
def seeded_sale(tenant, store, warehouse, owner):
    p = Product.objects.create(tenant=tenant, name="Café", price=Decimal("3000"), is_active=True)
    StockItem.objects.create(
        tenant=tenant, warehouse=warehouse, product=p,
        on_hand=Decimal("100"), avg_cost=Decimal("1000"), stock_value=Decimal("100000"),
    )
    create_sale(
        user=owner, tenant_id=tenant.id, store_id=store.id, warehouse_id=warehouse.id,
        lines_in=[{"product_id": p.id, "qty": "2", "unit_price": "3000"}],
        payments_in=[{"method": "cash", "amount": "6000"}],
    )
    return p


@pytest.mark.django_db
class TestSalesExport:
    def test_sales_excel_has_revenue_not_zero(self, api_client, seeded_sale, store):
        r = api_client.get("/api/reports/sales-summary/export/", HTTP_X_STORE_ID=str(store.id))
        assert r.status_code == 200, r.content[:200]
        rows = _xlsx_rows(r)
        # headers + al menos 1 día con datos
        assert len(rows) >= 2, rows
        data_rows = rows[1:]
        # Columna "Ingresos" (índice 2) debe ser > 0 (antes salía 0 por leer "total").
        total_ingresos = sum(float(rr[2] or 0) for rr in data_rows)
        assert total_ingresos > 0, f"Ingresos quedaron en 0 (regresión total→revenue): {data_rows}"


@pytest.mark.django_db
class TestStockValuedExport:
    def test_stock_excel_lists_product(self, api_client, seeded_sale, store):
        r = api_client.get("/api/reports/stock-valued/export/", HTTP_X_STORE_ID=str(store.id))
        assert r.status_code == 200, r.content[:200]
        rows = _xlsx_rows(r)
        assert len(rows) >= 2, f"Export de stock vacío (regresión warehouses→results): {rows}"
        # Producto (col 2) y Valor stock (col 5) presentes.
        names = [rr[2] for rr in rows[1:]]
        assert "Café" in names, names


@pytest.mark.django_db
class TestLossesExport:
    def test_losses_excel_has_value_and_sku(self, api_client, tenant, store, warehouse):
        p = Product.objects.create(tenant=tenant, name="Leche", sku="SKU-LECHE",
                                   price=Decimal("1000"), is_active=True)
        StockItem.objects.create(tenant=tenant, warehouse=warehouse, product=p,
                                 on_hand=Decimal("10"), avg_cost=Decimal("800"),
                                 stock_value=Decimal("8000"))
        StockMove.objects.create(
            tenant=tenant, warehouse=warehouse, product=p,
            move_type="OUT", ref_type="ISSUE", qty=Decimal("2"),
            cost_snapshot=Decimal("800"), value_delta=Decimal("-1600"),
            reason="MERMA", created_by=None,
        )
        r = api_client.get("/api/reports/losses/export/", HTTP_X_STORE_ID=str(store.id))
        assert r.status_code == 200, r.content[:200]
        rows = _xlsx_rows(r)
        assert len(rows) >= 2, f"Export de mermas vacío: {rows}"
        body = rows[1:]
        # SKU (col 2) y Costo perdido (col 6) deben venir poblados.
        assert any(rr[2] == "SKU-LECHE" for rr in body), body
        assert any(float(rr[6] or 0) > 0 for rr in body), f"value_lost quedó en 0: {body}"


@pytest.mark.django_db
class TestAuditTrailStoreScope:
    def test_export_only_active_store_moves(self, api_client, tenant, store, warehouse):
        # Tienda B con su propia bodega y movimiento.
        store_b = Store.objects.create(tenant=tenant, name="Local B")
        wh_b = Warehouse.objects.create(tenant=tenant, store=store_b, name="WH-B")
        p = Product.objects.create(tenant=tenant, name="ProdAudit", price=Decimal("1"), is_active=True)
        StockMove.objects.create(tenant=tenant, warehouse=warehouse, product=p,
                                 move_type="IN", ref_type="RECEIVE", qty=Decimal("5"),
                                 cost_snapshot=Decimal("0"), value_delta=Decimal("0"))
        StockMove.objects.create(tenant=tenant, warehouse=wh_b, product=p,
                                 move_type="IN", ref_type="RECEIVE", qty=Decimal("99"),
                                 cost_snapshot=Decimal("0"), value_delta=Decimal("0"))
        # Export con contexto de tienda A.
        r = api_client.get("/api/reports/audit-trail/export/", HTTP_X_STORE_ID=str(store.id))
        assert r.status_code == 200, r.content[:200]
        rows = _xlsx_rows(r)
        qtys = [float(rr[5] or 0) for rr in rows[1:]]
        assert 5.0 in qtys, f"falta el movimiento de la tienda activa: {rows}"
        assert 99.0 not in qtys, f"FUGA cross-store: apareció movimiento de tienda B: {rows}"


@pytest.mark.django_db
class TestAbcWeeklyEmail:
    def test_weekly_email_sends_after_fix(self, seeded_sale, owner, store):
        from django.core import mail
        from reports.tasks import send_weekly_abc_report
        owner.email = "dueno@marbrava.cl"
        owner.save(update_fields=["email"])
        # Asegurar store activa.
        store.is_active = True
        store.save(update_fields=["is_active"])

        mail.outbox = []
        send_weekly_abc_report()
        assert len(mail.outbox) == 1, (
            "el email semanal ABC no se envió (regresión items→results en tasks.py)")
        assert "dueno@marbrava.cl" in mail.outbox[0].to
